# modules/process_design/tabs/equipment_import_export.py
import os
import sys
import csv
import json
import re
import traceback
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QTimer, QPoint, QRect
from PySide6.QtWidgets import (
    QWidget, QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLineEdit, 
    QLabel, QMessageBox, QFileDialog, QCheckBox, QTextEdit,
    QGroupBox, QScrollArea, QDialogButtonBox, QComboBox
)

from .equipment_id_generator import EquipmentIDGenerator
from .equipment_templates import (
    EquipmentTemplateCreator,
    EquipmentTemplateFiller,
    TemplateTypeDialog
)


class EquipmentImportExport:
    """设备导入导出管理器"""
    
    def __init__(self, parent_tab):
        """
        初始化导入导出管理器
        
        参数:
            parent_tab: 父标签页对象
        """
        self.parent = parent_tab
        self.data_manager = parent_tab.data_manager
        self.process_manager = parent_tab.process_manager
        
    # ==================== 核心导入导出方法 ====================
    
    def import_equipment(self):
        """导入设备数据"""
        file_path, selected_filter = QFileDialog.getOpenFileName(
            self.parent, "导入设备数据", "", 
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;JSON文件 (*.json)"
        )
        
        if not file_path:
            return
        
        try:
            imported_count = 0
            
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                imported_count = self._import_excel_file(file_path)
            elif file_path.endswith('.csv'):
                imported_count = self._import_csv_file(file_path)
            elif file_path.endswith('.json'):
                imported_count = self._import_json_file(file_path)
            
            # 刷新设备列表
            self.parent.load_equipment()
            self.parent.equipment_list_updated.emit()
            
            QMessageBox.information(self.parent, "导入成功", 
                                  f"成功导入 {imported_count} 个设备")
            
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"导入错误详情: {error_details}")
            QMessageBox.critical(self.parent, "导入失败", 
                               f"导入文件时发生错误:\n{str(e)}")
    
    def _import_excel_file(self, file_path):
        """导入Excel文件"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 查找表头行
        header_row = self._find_header_row(ws)
        if header_row is None:
            header_row = 9  # 默认值
        
        print(f"表头行: {header_row}, 最大列数: {ws.max_column}")
        
        # 读取表头，建立列映射
        column_mapping = self._parse_excel_header(ws, header_row)
        print(f"列映射结果: {column_mapping}")
        
        # 检查必要字段
        if "unique_code" not in column_mapping.values():
            QMessageBox.warning(self.parent, "警告", 
                              "导入文件缺少'唯一编码'列，请使用正确的模板文件\n"
                              "或使用'修复文件'功能添加唯一编码")
            return 0
        
        # 读取数据行
        imported_count = 0
        data_start_row = header_row + 1
        
        for row in range(data_start_row, ws.max_row + 1):
            # 检查是否到达数据末尾
            name = self._get_cell_value_by_mapping(ws, row, column_mapping, 'name')
            if not name:
                continue
            
            # 读取唯一编码
            unique_code = self._get_cell_value_by_mapping(ws, row, column_mapping, 'unique_code')
            if not unique_code:
                self.parent.status_bar.setText(f"第{row}行缺少唯一编码，已跳过")
                continue
            
            # 验证唯一编码
            is_valid, message = EquipmentIDGenerator.validate_equipment_id(unique_code)
            if not is_valid:
                self.parent.status_bar.setText(f"第{row}行唯一编码验证失败: {message}")
                continue
            
            # 创建设备数据
            equipment_data = self._create_equipment_data_from_row(ws, row, column_mapping, unique_code)
            
            # 添加名称对照
            if equipment_data.get('name') and equipment_data.get('description_en'):
                self.data_manager.add_equipment_name_mapping(
                    equipment_data['name'], 
                    equipment_data['description_en']
                )
            
            # 保存设备
            if self._save_equipment_data(equipment_data):
                imported_count += 1
                self.parent.status_bar.setText(f"正在导入第{row}行数据...")
        
        return imported_count
    
    def _import_csv_file(self, file_path):
        """导入CSV文件"""
        imported_count = 0
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            csv_reader = csv.DictReader(f)
            
            for row in csv_reader:
                # 检查唯一编码
                if '唯一编码' not in row or not row['唯一编码']:
                    continue
                
                # 验证唯一编码
                unique_code = row['唯一编码'].strip()
                is_valid, message = EquipmentIDGenerator.validate_equipment_id(unique_code)
                if not is_valid:
                    continue
                
                # 创建设备数据
                equipment_data = self._create_equipment_data_from_csv_row(row)
                
                # 保存设备
                if self._save_equipment_data(equipment_data):
                    imported_count += 1
        
        return imported_count
    
    def _import_json_file(self, file_path):
        """导入JSON文件"""
        imported_count = 0
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
            for item in data:
                if 'unique_code' not in item or not item['unique_code']:
                    continue
                
                # 验证唯一编码
                is_valid, message = EquipmentIDGenerator.validate_equipment_id(item['unique_code'])
                if not is_valid:
                    continue
                
                # 保存设备
                if self._save_equipment_data(item):
                    imported_count += 1
        
        return imported_count
    
    # ==================== 辅助方法 ====================
    
    def _find_header_row(self, ws):
        """查找表头行"""
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_text = str(cell_value)
                cleaned_text = cell_text.replace('<br>', '\n').replace('<BR>', '\n').strip()
                
                if ('序号' in cleaned_text or 
                    'Item' in cleaned_text or 
                    '设备位号' in cleaned_text or
                    cleaned_text in ['Item\n序号', 'Item<br>序号', '序号', 'Item']):
                    print(f"找到表头行: 第{row}行, 内容: '{cell_text}'")
                    return row
        
        # 尝试其他方式查找
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '设备位号' in str(cell_value):
                    print(f"通过设备位号找到表头行: 第{row}行")
                    return row
        
        return None
    
    def _parse_excel_header(self, ws, header_row):
        """解析Excel表头"""
        column_mapping = {}
        max_col = ws.max_column
        
        for col in range(1, max_col + 1):
            header_cell = ws.cell(row=header_row, column=col)
            if header_cell.value:
                header_text = str(header_cell.value)
                normalized_text = header_text.replace('<br>', '\n').replace('<BR>', '\n').strip()
                
                print(f"列{col}表头: '{header_text}' -> 规范化: '{normalized_text}'")
                
                field_name = self._map_header_to_field(normalized_text)
                if field_name:
                    column_mapping[col] = field_name
        
        return column_mapping
    
    def _map_header_to_field(self, header_text):
        """映射表头到字段名"""
        mappings = [
            ("Item\n序号", None),  # 跳过序号列
            ("Tag num.\n设备位号", "equipment_id"),
            ("Description", "description_en"),
            ("设备名称", "name"),
            ("P&ID DWG. NO.\nP&ID图号", "pid_dwg_no"),
            ("Technical specifications\n设备技术规格", "specification"),
            ("QTY.\n数量", "quantity"),
            ("Unit price\n单价\nCNY", "unit_price"),
            ("Total price\n总价\nCNY", "total_price"),
            ("Design tem.\n设计温度\n℃", "design_temperature"),
            ("Design pressure\n设计压力\nMPa·G", "design_pressure"),
            ("Operating tem.\n操作温度\n℃", "operating_temperature"),
            ("Operating pressure\n操作压力\nMPa·G", "operating_pressure"),
            ("Unit power\n单机功率\nkW", "single_power"),
            ("Run power\n运行功率\nkW", "operating_power"),
            ("Installed power\n装机功率\nkW", "total_power"),
            ("Material\n材质", "material"),
            ("Insulation\n保温", "insulation"),
            ("Dry weight\n单机重量\nt", "weight_estimate"),
            ("Operating weight\n操作重量\nt", "operating_weight"),
            ("Total weight\总重量\nt", "total_weight"),
            ("Dynamic\n荷载系数", "dynamic"),
            ("Remark\n备注", "notes"),
            ("唯一编码", "unique_code")
        ]
        
        for header_pattern, field in mappings:
            if self._matches_header(header_text, header_pattern):
                return field
        
        return None
    
    def _matches_header(self, header_text, expected):
        """检查表头是否匹配"""
        header_text = str(header_text).strip()
        expected_text = str(expected).strip()
        
        # 移除换行符和特殊字符
        normalized_header = header_text.replace('\n', ' ').replace('<br>', ' ').strip()
        normalized_expected = expected_text.replace('\n', ' ').strip()
        
        # 直接比较
        if normalized_header == normalized_expected:
            return True
        
        # 检查是否包含关键词
        keywords = normalized_expected.split()
        for keyword in keywords:
            if keyword.lower() in header_text.lower():
                return True
        
        # 检查清理后的文本
        clean_header = re.sub(r'[^\w\u4e00-\u9fff]', '', header_text)
        for keyword in keywords:
            clean_keyword = re.sub(r'[^\w\u4e00-\u9fff]', '', keyword)
            if clean_keyword and clean_keyword in clean_header:
                return True
        
        return False
    
    def _get_cell_value_by_mapping(self, ws, row, column_mapping, field_name, default=''):
        """根据字段映射获取单元格值"""
        for col, field in column_mapping.items():
            if field == field_name:
                cell_value = ws.cell(row=row, column=col).value
                return str(cell_value).strip() if cell_value is not None else default
        return default
    
    def _create_equipment_data_from_row(self, ws, row, column_mapping, unique_code):
        """从Excel行创建设备数据"""
        equipment_data = {"unique_code": unique_code}
        
        # 生成或获取设备ID
        equipment_id = self._get_cell_value_by_mapping(ws, row, column_mapping, 'equipment_id')
        if not equipment_id:
            equipment_name = self._get_cell_value_by_mapping(ws, row, column_mapping, 'name')
            if equipment_name:
                base_id = equipment_name.replace(" ", "_")
                equipment_id = f"EQ-{base_id}"
        
        equipment_data["equipment_id"] = equipment_id
        
        # 读取其他字段
        field_processors = {
            'name': lambda v: str(v).strip() if v else "",
            'description_en': lambda v: str(v).strip() if v else "",
            'specification': lambda v: str(v).strip() if v else "",
            'pid_dwg_no': lambda v: str(v).strip() if v else "",
            'quantity': self._parse_integer,
            'unit_price': self._parse_float,
            'total_price': self._parse_float,
            'design_temperature': self._parse_temperature,
            'design_pressure': self._parse_pressure,
            'operating_temperature': self._parse_temperature,
            'operating_pressure': self._parse_pressure,
            'single_power': self._parse_float,
            'operating_power': self._parse_float,
            'total_power': self._parse_float,
            'material': lambda v: str(v).strip() if v else "",
            'insulation': lambda v: str(v).strip() if v else "",
            'weight_estimate': self._parse_float,
            'operating_weight': self._parse_float,
            'total_weight': self._parse_float,
            'dynamic': lambda v: str(v).strip() if v else "",
            'notes': lambda v: str(v).strip() if v else "",
        }
        
        for col, field in column_mapping.items():
            if field in ['unique_code', 'equipment_id']:
                continue
            
            cell_value = ws.cell(row=row, column=col).value
            if field in field_processors:
                equipment_data[field] = field_processors[field](cell_value)
            else:
                equipment_data[field] = str(cell_value).strip() if cell_value else ""
        
        # 设置默认值
        if 'type' not in equipment_data:
            equipment_data['type'] = '其他'
        if 'status' not in equipment_data:
            equipment_data['status'] = '运行中'
        
        return equipment_data
    
    def _create_equipment_data_from_csv_row(self, row):
        """从CSV行创建设备数据"""
        equipment_data = {}
        
        field_mapping = {
            '唯一编码': 'unique_code',
            '设备位号': 'equipment_id',
            '设备名称': 'name',
            '英文描述': 'description_en',
            '规格': 'specification',
            '数量': 'quantity',
            '单价': 'unit_price',
            '总价': 'total_price',
            '设计温度': 'design_temperature',
            '设计压力': 'design_pressure',
            '操作温度': 'operating_temperature',
            '操作压力': 'operating_pressure',
            '单机功率': 'single_power',
            '运行功率': 'operating_power',
            '装机功率': 'total_power',
            '材质': 'material',
            '保温': 'insulation',
            '单机重量': 'weight_estimate',
            '操作重量': 'operating_weight',
            '总重量': 'total_weight',
            '荷载系数': 'dynamic',
            '备注': 'notes',
            'P&ID图号': 'pid_dwg_no'
        }
        
        for csv_field, data_field in field_mapping.items():
            if csv_field in row:
                value = row[csv_field].strip()
                if value:
                    # 根据字段类型处理值
                    if data_field in ['quantity']:
                        equipment_data[data_field] = self._parse_integer(value)
                    elif data_field in ['unit_price', 'total_price', 'single_power', 
                                       'operating_power', 'total_power', 'weight_estimate',
                                       'operating_weight', 'total_weight']:
                        equipment_data[data_field] = self._parse_float(value)
                    elif data_field in ['design_temperature', 'operating_temperature']:
                        equipment_data[data_field] = self._parse_temperature(value)
                    elif data_field in ['design_pressure', 'operating_pressure']:
                        equipment_data[data_field] = self._parse_pressure(value)
                    else:
                        equipment_data[data_field] = value
        
        # 设置默认值
        if 'type' not in equipment_data:
            equipment_data['type'] = '其他'
        if 'status' not in equipment_data:
            equipment_data['status'] = '运行中'
        
        return equipment_data
    
    def _parse_integer(self, value):
        """解析整数"""
        if value is None or value == '':
            return 1
        
        try:
            if isinstance(value, (int, float)):
                return int(value)
            else:
                return int(float(str(value)))
        except (ValueError, TypeError):
            return 1
    
    def _parse_float(self, value):
        """解析浮点数"""
        if value is None or value == '':
            return None
        
        try:
            if isinstance(value, (int, float)):
                return float(value)
            else:
                value_str = str(value).strip()
                if value_str and value_str.upper() not in ['NT', 'NP']:
                    return float(value_str)
                else:
                    return None
        except (ValueError, TypeError):
            return None
    
    def _parse_temperature(self, value):
        """解析温度值"""
        if value is None or value == '':
            return None
        
        if isinstance(value, str) and value.upper() == 'NT':
            return 'NT'
        
        return self._parse_float(value)
    
    def _parse_pressure(self, value):
        """解析压力值"""
        if value is None or value == '':
            return None
        
        if isinstance(value, str) and value.upper() == 'NP':
            return 'NP'
        
        return self._parse_float(value)
    
    def _save_equipment_data(self, equipment_data):
        """保存设备数据"""
        try:
            from ..data.data_models import UnifiedEquipment
            
            # 使用from_dict方法，它会自动进行安全转换
            equipment = UnifiedEquipment.from_dict(equipment_data)
            
            if self.process_manager.add_equipment(equipment):
                print(f"✅ 设备数据保存成功: {equipment.name}")
                return True
            else:
                print(f"❌ 设备数据保存失败: {equipment.name}")
                return False
                
        except Exception as e:
            print(f"❌ 保存设备数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    # ==================== 导出功能 ====================
    
    def export_equipment(self):
        """导出设备数据（按照设备清单模板格式）"""
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self.parent, "导出设备清单", "设备清单.xlsx",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv);;JSON文件 (*.json)"
        )
        
        if not file_path:
            return
        
        try:
            # 获取所有设备
            equipment_list = self.process_manager.get_all_equipment()
            
            # 按照模板格式构建数据
            rows = []
            for i, equipment in enumerate(equipment_list, 1):
                row = self._create_export_row(equipment, i)
                rows.append(row)
            
            print(f"总共导出 {len(rows)} 行数据")
            if rows:
                print("第一行数据示例:")
                for key, value in rows[0].items():
                    print(f"  {key}: {value}")

            # 创建DataFrame
            df = pd.DataFrame(rows)
            
            # 设置列顺序（按照中文模板顺序）
            columns = [
                'Item\n序号', 
                'Tag num.\n设备位号', 
                'Description', 
                '设备名称', 
                'P&ID DWG. NO.\nP&ID图号', 
                'Technical specifications\n设备技术规格', 
                'QTY.\n数量', 
                'Unit price\n单价\nCNY', 
                'Total price\n总价\nCNY',
                'Design tem.\n设计温度\n℃', 
                'Design pressure\n设计压力\nMPa·G', 
                'Operating tem.\n操作温度\n℃', 
                'Operating pressure\n操作压力\nMPa·G', 
                'Unit power\n单机功率\nkW', 
                'Run power\n运行功率\nkW', 
                'Installed power\n装机功率\nkW',
                'Material\n材质', 
                'Insulation\n保温', 
                'Dry weight\n单机重量\nt', 
                'Dynamic\n荷载系数', 
                'Remark\n备注', 
                '唯一编码'
            ]
            
            # 重新排列列顺序
            df = df.reindex(columns=[col for col in columns if col in df.columns])
            
            # 根据文件类型保存
            if file_path.endswith('.xlsx'):
                self._export_to_excel(df, file_path)
            elif file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
            elif file_path.endswith('.json'):
                df.to_json(file_path, orient='records', force_ascii=False, indent=2)
            
            QMessageBox.information(self.parent, "导出成功", 
                                  f"成功导出 {len(equipment_list)} 个设备\n文件已保存到: {file_path}")
            
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"导出错误详情: {error_details}")
            QMessageBox.critical(self.parent, "导出失败", 
                               f"导出文件时发生错误:\n{str(e)}")
    
    def _create_export_row(self, equipment, index):
        """创建导出数据行"""
        # 获取英文描述
        description_en = ""
        if hasattr(equipment, 'description_en') and equipment.description_en:
            description_en = equipment.description_en
        elif equipment.name:
            # 尝试从对照表获取
            description_en = self.data_manager.get_english_name(equipment.name)
        
        item_number = f"{index:02d}"
        
        # 确保设备ID不为空
        equipment_id = equipment.equipment_id
        if not equipment_id:
            # 如果设备ID为空，使用设备名称生成
            base_id = equipment.name.replace(" ", "_") if equipment.name else f"EQUIP_{index}"
            equipment_id = f"EQ-{base_id}"
        
        # 处理设计温度和压力值
        design_temp = equipment.design_temperature
        if isinstance(design_temp, (int, float)):
            design_temp_str = f"{design_temp}"
        else:
            design_temp_str = str(design_temp) if design_temp else ""
        
        design_pressure = equipment.design_pressure
        if isinstance(design_pressure, (int, float)):
            design_pressure_str = f"{design_pressure}"
        else:
            design_pressure_str = str(design_pressure) if design_pressure else ""
        
        # 处理数量字段，确保是整数
        quantity = equipment.quantity if hasattr(equipment, 'quantity') else 1
        try:
            quantity = int(float(quantity))
        except (ValueError, TypeError):
            quantity = 1
        
        return {
            'Item\n序号': item_number,
            'Tag num.\n设备位号': equipment_id,
            'Description': description_en,
            '设备名称': equipment.name,
            'P&ID DWG. NO.\nP&ID图号': equipment.pid_dwg_no if hasattr(equipment, 'pid_dwg_no') else '',
            'Technical specifications\n设备技术规格': equipment.specification,
            'QTY.\n数量': quantity,
            'Unit price\n单价\nCNY': equipment.unit_price if hasattr(equipment, 'unit_price') else '',
            'Total price\n总价\nCNY': equipment.total_price if hasattr(equipment, 'total_price') else '',
            'Design tem.\n设计温度\n℃': design_temp_str,
            'Design pressure\n设计压力\nMPa·G': design_pressure_str,
            'Operating tem.\n操作温度\n℃': equipment.operating_temperature if hasattr(equipment, 'operating_temperature') else '',
            'Operating pressure\n操作压力\nMPa·G': equipment.operating_pressure if hasattr(equipment, 'operating_pressure') else '',
            'Unit power\n单机功率\nkW': equipment.single_power if hasattr(equipment, 'single_power') else '',
            'Run power\n运行功率\nkW': equipment.operating_power if hasattr(equipment, 'operating_power') else '',
            'Installed power\n装机功率\nkW': equipment.total_power if hasattr(equipment, 'total_power') else '',
            'Material\n材质': getattr(equipment, 'material', ''),
            'Insulation\n保温': equipment.insulation if hasattr(equipment, 'insulation') else '',
            'Dry weight\n单机重量\nt': equipment.weight_estimate if hasattr(equipment, 'weight_estimate') else '',
            'Dynamic\n荷载系数': equipment.dynamic if hasattr(equipment, 'dynamic') else '',
            'Remark\n备注': equipment.notes if hasattr(equipment, 'notes') else '',
            '唯一编码': equipment.unique_code if hasattr(equipment, 'unique_code') else EquipmentIDGenerator.generate_equipment_id(
                equipment.equipment_type if hasattr(equipment, 'equipment_type') else equipment.type,
                equipment.name
            )
        }
    
    def _export_to_excel(self, df, file_path):
        """导出到Excel文件"""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='设备清单')
            
            # 获取工作表对象以设置格式
            worksheet = writer.sheets['设备清单']
            
            # 设置列宽
            column_widths = {
                'Item\n序号': 7,
                'Tag num.\n设备位号': 22,
                'Description': 45,
                '设备名称': 28,
                'P&ID DWG. NO.\nP&ID图号': 35,
                'Technical specifications\n设备技术规格': 60,
                'QTY.\n数量': 10,
                'Unit price\n单价\nCNY': 10,
                'Total price\n总价\nCNY': 10,
                'Design tem.\n设计温度\n℃': 10,
                'Design pressure\n设计压力\nMPa·G': 10,
                'Operating tem.\n操作温度\n℃': 10,
                'Operating pressure\n操作压力\nMPa·G': 10,
                'Unit power\n单机功率\nkW': 10,
                'Run power\n运行功率\nkW': 10,
                'Installed power\n装机功率\nkW': 10,
                'Material\n材质': 10,
                'Insulation\n保温': 10,
                'Dry weight\n单机重量\nt': 10,
                'Dynamic\n荷载系数': 10,
                'Remark\n备注': 20,
                '唯一编码': 30
            }
            
            for i, col in enumerate(df.columns, 1):
                width = column_widths.get(col, 15)
                worksheet.column_dimensions[get_column_letter(i)].width = width
            
            # 设置标题行样式
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # ==================== 文件修复功能 ====================
    
    def repair_import_file(self):
        """修复导入文件格式 - 主要修复唯一编码和设备位号"""
        file_path, _ = QFileDialog.getOpenFileName(
            self.parent, "选择要修复的文件", "", 
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            # 加载文件
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 查找表头行
            header_row = None
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        cell_text = str(cell_value)
                        if any(keyword in cell_text for keyword in ["序号", "Item", "设备位号", "Tag num", "唯一编码"]):
                            header_row = row
                            break
                if header_row:
                    break
            
            if header_row is None:
                header_row = 1
            
            # 建立列映射
            column_mapping = {}
            has_unique_code = False
            has_equipment_id = False
            has_name = False
            
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value:
                    header_text = str(header_cell.value).strip()
                    
                    if self._matches_header(header_text, "唯一编码", ["唯一编码", "unique_code", "Unique Code"]):
                        column_mapping["unique_code"] = col
                        has_unique_code = True
                    elif self._matches_header(header_text, "Tag num.\n设备位号", ["Tag num", "设备位号", "equipment_id", "Equipment ID"]):
                        column_mapping["equipment_id"] = col
                        has_equipment_id = True
                    elif self._matches_header(header_text, "设备名称", ["设备名称", "name", "Name", "Description", "description"]):
                        column_mapping["name"] = col
                        has_name = True
                    elif self._matches_header(header_text, "设备类型", ["设备类型", "type", "Type", "equipment_type"]):
                        column_mapping["type"] = col
            
            # 如果没有设备名称列，尝试其他常见名称
            if not has_name:
                for col in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=header_row, column=col)
                    if header_cell.value:
                        header_text = str(header_cell.value).strip()
                        if any(keyword in header_text for keyword in ["名称", "Description", "description", "Desc"]):
                            column_mapping["name"] = col
                            has_name = True
                            break
            
            # 添加缺失的列
            new_col = ws.max_column + 1
            
            if not has_unique_code:
                ws.cell(row=header_row, column=new_col, value="唯一编码")
                column_mapping["unique_code"] = new_col
                new_col += 1
            
            if not has_equipment_id:
                ws.cell(row=header_row, column=new_col, value="设备位号")
                column_mapping["equipment_id"] = new_col
            
            # 修复数据行
            fixed_unique_code_count = 0
            fixed_equipment_id_count = 0
            skipped_rows = 0
            
            data_start_row = header_row + 1
            
            for row in range(data_start_row, ws.max_row + 1):
                # 检查是否是空行
                name_col = column_mapping.get("name")
                if name_col:
                    name_cell = ws.cell(row=row, column=name_col)
                    if not name_cell.value or str(name_cell.value).strip() == "":
                        skipped_rows += 1
                        continue
                
                # 修复唯一编码
                unique_code_col = column_mapping.get("unique_code")
                if unique_code_col:
                    unique_code_cell = ws.cell(row=row, column=unique_code_col)
                    current_unique_code = unique_code_cell.value
                    
                    if not current_unique_code or not str(current_unique_code).strip():
                        equipment_name = ""
                        if "name" in column_mapping:
                            name_cell = ws.cell(row=row, column=column_mapping["name"])
                            equipment_name = str(name_cell.value).strip() if name_cell.value else ""
                        
                        equipment_type = "其他"
                        if "type" in column_mapping:
                            type_cell = ws.cell(row=row, column=column_mapping["type"])
                            if type_cell.value:
                                equipment_type = str(type_cell.value).strip()
                        
                        unique_code = EquipmentIDGenerator.generate_equipment_id(
                            equipment_type, 
                            equipment_name
                        )
                        
                        unique_code_cell.value = unique_code
                        fixed_unique_code_count += 1
                
                # 修复设备位号
                equipment_id_col = column_mapping.get("equipment_id")
                if equipment_id_col:
                    equipment_id_cell = ws.cell(row=row, column=equipment_id_col)
                    current_equipment_id = equipment_id_cell.value
                    
                    if not current_equipment_id or not str(current_equipment_id).strip():
                        equipment_name = ""
                        if "name" in column_mapping:
                            name_cell = ws.cell(row=row, column=column_mapping["name"])
                            if name_cell.value:
                                equipment_name = str(name_cell.value).strip()
                        
                        if equipment_name:
                            base_id = equipment_name.replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_")
                            equipment_id = f"EQ-{base_id}"
                            
                            if len(equipment_id) > 50:
                                equipment_id = equipment_id[:50]
                        else:
                            equipment_id = f"EQ-UNKNOWN-{row - data_start_row + 1:04d}"
                        
                        equipment_id_cell.value = equipment_id
                        fixed_equipment_id_count += 1
            
            # 生成保存路径
            import os
            base_name = os.path.splitext(file_path)[0]
            save_path = f"{base_name}_修复后.xlsx"
            
            # 避免覆盖已存在的文件
            counter = 1
            original_save_path = save_path
            while os.path.exists(save_path):
                name_without_ext, _ = os.path.splitext(original_save_path)
                save_path = f"{name_without_ext}_{counter}.xlsx"
                counter += 1
            
            # 保存工作簿
            wb.save(save_path)
            
            # 显示修复结果
            self._show_repair_result(fixed_unique_code_count, fixed_equipment_id_count, 
                                   skipped_rows, save_path)
            
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"修复文件时出错: {error_details}")
            QMessageBox.critical(self.parent, "修复失败", 
                               f"修复文件时发生错误:\n{str(e)}")
    
    def _show_repair_result(self, fixed_unique_code, fixed_equipment_id, skipped, save_path):
        """显示修复结果对话框"""
        result_message = f"文件修复完成！\n\n"
        result_message += f"修复统计：\n"
        result_message += f"- 生成唯一编码: {fixed_unique_code} 个\n"
        result_message += f"- 生成设备位号: {fixed_equipment_id} 个\n"
        result_message += f"- 跳过空行: {skipped} 行\n\n"
        result_message += f"修复后的文件已保存到：\n{save_path}"
        
        # 创建详细结果对话框
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("修复完成")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # 结果信息
        result_label = QLabel(result_message)
        result_label.setWordWrap(True)
        layout.addWidget(result_label)
        
        # 添加打开文件按钮
        btn_layout = QHBoxLayout()
        
        open_btn = QPushButton("打开修复后的文件")
        open_btn.clicked.connect(lambda: self.parent.open_file(save_path))
        btn_layout.addWidget(open_btn)
        
        import os
        open_folder_btn = QPushButton("打开所在文件夹")
        open_folder_btn.clicked.connect(lambda: self.parent.open_file(os.path.dirname(save_path)))
        btn_layout.addWidget(open_folder_btn)
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    # ==================== 模板相关方法 ====================
    
    def import_equipment_by_template(self):
        """模板导入：导入符合模板格式的Excel文件"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self.parent, "选择模板文件导入", "",
                "Excel模板文件 (*.xlsx);;所有文件 (*)"
            )
            
            if not file_path:
                return
            
            # 询问是否导入项目信息
            reply = QMessageBox.question(
                self.parent, "导入项目信息",
                "是否同时导入文件中的项目信息？\n"
                "（项目名称、子项名称、文件编号等）",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Cancel:
                return
            
            import_project_info = (reply == QMessageBox.Yes)
            
            # 解析模板文件
            result = self._parse_template_file(file_path, import_project_info)
            
            if not result:
                QMessageBox.warning(self.parent, "导入失败", 
                                  "文件解析失败，请检查文件格式")
                return
            
            project_info, equipment_list = result
            
            if not equipment_list:
                QMessageBox.warning(self.parent, "警告", 
                                  "文件中没有找到设备数据")
                return
            
            # 显示预览对话框
            dialog = TemplateImportPreviewDialog(project_info, equipment_list, self.parent)
            
            if dialog.exec() == QDialog.Accepted:
                # 获取用户选择的导入选项
                import_options = dialog.get_import_options()
                
                # 执行导入
                success_count = self._execute_template_import(
                    equipment_list, 
                    import_options
                )
                
                if success_count > 0:
                    QMessageBox.information(
                        self.parent, "导入成功",
                        f"成功导入 {success_count} 个设备"
                    )
                    self.parent.load_equipment()
                    self.parent.equipment_list_updated.emit()
                else:
                    QMessageBox.warning(self.parent, "导入失败", 
                                      "没有设备被导入")
            
        except Exception as e:
            QMessageBox.critical(self.parent, "导入错误", 
                               f"导入过程中发生错误:\n{str(e)}")
            traceback.print_exc()
    
    def _parse_template_file(self, file_path, import_project_info=True):
        """解析模板格式的Excel文件"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            project_info = {}
            equipment_list = []
            
            # ==================== 解析项目信息 ====================
            if import_project_info:
                project_info = self._parse_project_info_from_sheet(ws)
            
            # ==================== 查找表头行 ====================
            header_row = self._find_template_header_row(ws)
            
            if header_row is None:
                raise ValueError("无法找到表头行")
            
            # ==================== 解析表头映射 ====================
            column_mapping = self._parse_template_header_mapping(ws, header_row)
            
            # ==================== 解析设备数据 ====================
            data_start_row = header_row + 1
            
            for row in range(data_start_row, ws.max_row + 1):
                # 检查是否有效行
                name = self._get_cell_value_by_mapping(ws, row, column_mapping, 'name')
                if not name or name.strip() == '':
                    continue
                
                # 检查唯一编码
                unique_code = self._get_cell_value_by_mapping(ws, row, column_mapping, 'unique_code')
                if not unique_code or unique_code.strip() == '':
                    # 如果没有唯一编码，生成一个
                    equipment_id = self._get_cell_value_by_mapping(ws, row, column_mapping, 'equipment_id') or name
                    equipment_type = '其他'  # 默认类型，可以从其他地方推断
                    unique_code = EquipmentIDGenerator.generate_equipment_id(equipment_type, name)
                
                # 创建设备数据字典
                equipment_data = self._create_equipment_data_from_template_row(
                    ws, row, column_mapping, name, unique_code
                )
                
                equipment_list.append(equipment_data)
            
            return project_info, equipment_list
            
        except Exception as e:
            print(f"解析模板文件时出错: {e}")
            traceback.print_exc()
            return None
    
    def _find_template_header_row(self, ws):
        """查找模板文件表头行"""
        for row in range(1, 50):  # 在前50行查找表头
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and ('Item' in str(cell_value) or '序号' in str(cell_value)):
                return row
        
        # 尝试其他方式查找
        for row in range(1, 50):
            for col in range(1, 10):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '设备位号' in str(cell_value):
                    return row
        
        return None
    
    def _parse_template_header_mapping(self, ws, header_row):
        """解析模板文件表头映射"""
        column_mapping = {}
        
        # 预定义的表头映射（与模板导出格式一致）
        header_patterns = [
            ("Item\n序号", None),  # 忽略序号列
            ("Tag num.\n设备位号", "equipment_id"),
            ("Description", "description_en"),
            ("设备名称", "name"),
            ("P&ID DWG. NO.\nP&ID图号", "pid_dwg_no"),
            ("Technical specifications\n设备技术规格", "specification"),
            ("QTY.\n数量", "quantity"),
            ("Unit price\n单价\nCNY", "unit_price"),
            ("Total price\n总价\nCNY", "total_price"),
            ("Design tem.\n设计温度\n℃", "design_temperature"),
            ("Design pressure\n设计压力\nMPa·G", "design_pressure"),
            ("Operating tem.\n操作温度\n℃", "operating_temperature"),
            ("Operating pressure\n操作压力\nMPa·G", "operating_pressure"),
            ("Unit power\n单机功率\nkW", "single_power"),
            ("Run power\n运行功率\nkW", "operating_power"),
            ("Installed power\n装机功率\nkW", "total_power"),
            ("Material\n材质", "material"),
            ("Insulation\n保温", "insulation"),
            ("Dry weight\n单机重量\nt", "weight_estimate"),
            ("Operating weight\n操作重量\nt", "operating_weight"),
            ("Total weight\总重量\nt", "total_weight"),
            ("Dynamic\n荷载系数", "dynamic"),
            ("Remark\n备注", "notes"),
            ("唯一编码", "unique_code")
        ]
        
        # 遍历列，建立映射
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if not cell_value:
                continue
            
            header_text = str(cell_value)
            found = False
            
            for pattern, field in header_patterns:
                if self._normalize_header(header_text) == self._normalize_header(pattern):
                    if field:  # 跳过None的字段（如序号）
                        column_mapping[col] = field
                    found = True
                    break
            
            if not found:
                # 尝试部分匹配
                for pattern, field in header_patterns:
                    pattern_clean = self._normalize_header(pattern)
                    header_clean = self._normalize_header(header_text)
                    
                    if pattern and any(keyword in header_clean for keyword in pattern_clean.split()):
                        if field:
                            column_mapping[col] = field
                        break
        
        return column_mapping
    
    def _normalize_header(self, header_text):
        """规范化表头文本"""
        if not header_text:
            return ""
        
        # 替换所有换行符为空格
        text = str(header_text).replace('\n', ' ').replace('<br>', ' ').replace('<BR>', ' ')
        
        # 移除多余空格
        text = ' '.join(text.split())
        
        # 移除特殊字符
        text = re.sub(r'[^\w\u4e00-\u9fff\s]', '', text)
        
        return text.strip().lower()
    
    def _parse_project_info_from_sheet(self, ws):
        """从工作表中解析项目信息"""
        project_info = {}
        
        try:
            # 项目名称（A1单元格）
            if ws['A1'].value:
                project_info['project_name'] = str(ws['A1'].value).strip()
            
            # 子项名称（D1单元格）
            if ws['D1'].value:
                project_info['sub_item_name'] = str(ws['D1'].value).strip()
            
            # 文件编号（Q2单元格）
            if ws['Q2'].value:
                project_info['doc_no'] = str(ws['Q2'].value).strip()
            
            # 专业（E3单元格）
            if ws['E3'].value:
                project_info['speciality'] = str(ws['E3'].value).strip()
            
            # 阶段（E4单元格）
            if ws['E4'].value:
                project_info['phase'] = str(ws['E4'].value).strip()
            
            # 版次（S3单元格）
            if ws['S3'].value:
                project_info['revision'] = str(ws['S3'].value).strip()
            
            # 日期（V3单元格）
            if ws['V3'].value:
                project_info['date'] = str(ws['V3'].value).strip()
            
            # 导出日期（X3单元格）
            if ws['X3'].value:
                project_info['export_date'] = str(ws['X3'].value).strip()
                
        except Exception as e:
            print(f"解析项目信息时出错: {e}")
        
        return project_info
    
    def _create_equipment_data_from_template_row(self, ws, row, column_mapping, name, unique_code):
        """从模板行创建设备数据"""
        equipment_data = {
            'unique_code': unique_code.strip(),
            'equipment_id': self._get_cell_value_by_mapping(ws, row, column_mapping, 'equipment_id', default=''),
            'name': name.strip(),
            'description_en': self._get_cell_value_by_mapping(ws, row, column_mapping, 'description_en', default=''),
            'type': '其他',  # 从规格或其他地方推断
            'specification': self._get_cell_value_by_mapping(ws, row, column_mapping, 'specification', default=''),
            'pid_dwg_no': self._get_cell_value_by_mapping(ws, row, column_mapping, 'pid_dwg_no', default=''),
            'quantity': self._parse_integer(self._get_cell_value_by_mapping(ws, row, column_mapping, 'quantity', default=1)),
            'unit_price': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'unit_price')),
            'total_price': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'total_price')),
            'design_temperature': self._parse_temperature(self._get_cell_value_by_mapping(ws, row, column_mapping, 'design_temperature')),
            'design_pressure': self._parse_pressure(self._get_cell_value_by_mapping(ws, row, column_mapping, 'design_pressure')),
            'operating_temperature': self._parse_temperature(self._get_cell_value_by_mapping(ws, row, column_mapping, 'operating_temperature')),
            'operating_pressure': self._parse_pressure(self._get_cell_value_by_mapping(ws, row, column_mapping, 'operating_pressure')),
            'single_power': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'single_power')),
            'operating_power': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'operating_power')),
            'total_power': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'total_power')),
            'material': self._get_cell_value_by_mapping(ws, row, column_mapping, 'material', default=''),
            'insulation': self._get_cell_value_by_mapping(ws, row, column_mapping, 'insulation', default=''),
            'weight_estimate': self._parse_float(self._get_cell_value_by_mapping(ws, row, column_mapping, 'weight_estimate')),
            'dynamic': self._get_cell_value_by_mapping(ws, row, column_mapping, 'dynamic', default=''),
            'notes': self._get_cell_value_by_mapping(ws, row, column_mapping, 'notes', default=''),
            'status': '运行中',
            'manufacturer': '',
            'model': '',
            'location': ''
        }
        
        # 从规格中推断设备类型
        spec = equipment_data['specification']
        if spec:
            if '反应器' in spec or '反应釜' in spec:
                equipment_data['type'] = '反应器'
            elif '储罐' in spec or '罐' in spec:
                equipment_data['type'] = '储罐'
            elif '泵' in spec:
                equipment_data['type'] = '泵'
            elif '换热器' in spec:
                equipment_data['type'] = '换热器'
        
        # 如果设备ID为空，从设备名称生成
        if not equipment_data['equipment_id']:
            base_id = equipment_data['name'].replace(' ', '_').replace('(', '').replace(')', '')
            equipment_data['equipment_id'] = f"EQ-{base_id}"
        
        return equipment_data
    
    def _execute_template_import(self, equipment_list, import_options):
        """执行模板导入"""
        success_count = 0
        
        for equipment_data in equipment_list:
            try:
                from ..data.data_models import UnifiedEquipment
                
                # 根据导入选项处理数据
                if import_options.get('skip_existing', False):
                    # 检查设备是否已存在
                    existing = self.process_manager.get_equipment(equipment_data['equipment_id'])
                    if existing:
                        continue
                
                # 创建设备对象
                equipment = UnifiedEquipment(**equipment_data)
                
                # 保存设备
                if self.process_manager.add_equipment(equipment):
                    success_count += 1
                    
                    # 如果需要，添加名称对照
                    if equipment_data.get('name') and equipment_data.get('description_en'):
                        self.data_manager.add_equipment_name_mapping(
                            equipment_data['name'], 
                            equipment_data['description_en']
                        )
            
            except Exception as e:
                print(f"导入设备 {equipment_data.get('name', '未知')} 时出错: {e}")
                continue
        
        return success_count
    
    def export_equipment_with_template(self):
        """统一的设备导出功能 - 支持模板导出和批量导出"""
        try:
            # 检查是否有选中设备
            selected_ids = self.parent.get_selected_equipment_ids()
            
            # 确定导出范围
            if selected_ids:
                # 有选中设备，询问用户导出范围
                scope = self._ask_export_scope(len(selected_ids))
                if scope == "cancel":
                    return
                
                if scope == "selected":
                    # 导出选中设备
                    equipment_list = []
                    for equipment_id in selected_ids:
                        equipment = self.process_manager.get_equipment(equipment_id)
                        if equipment:
                            equipment_list.append(equipment)
                    export_scope = f"选中设备 ({len(selected_ids)}个)"
                else:  # all
                    # 导出全部设备
                    equipment_list = self.process_manager.get_all_equipment()
                    export_scope = "全部设备"
            else:
                # 没有选中设备，直接导出全部
                equipment_list = self.process_manager.get_all_equipment()
                export_scope = "全部设备"
            
            if not equipment_list:
                QMessageBox.warning(self.parent, "警告", "没有设备可导出")
                return
            
            # 选择模板文件
            template_path, _ = QFileDialog.getOpenFileName(
                self.parent, f"选择模板文件 - {export_scope}", "",
                "Excel模板文件 (*.xlsx);;所有文件 (*)"
            )
            
            if not template_path:
                return
            
            # 选择保存位置
            default_name = f"设备清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path, _ = QFileDialog.getSaveFileName(
                self.parent, f"保存设备清单 - {export_scope}", default_name,
                "Excel文件 (*.xlsx)"
            )
            
            if not output_path:
                return
            
            template_filler = EquipmentTemplateFiller()
            
            # 获取项目信息
            project_info = self._get_project_info_from_dialog()
            if project_info is None:
                return
            
            # 填充模板
            success = template_filler.fill_template(
                template_path, output_path, equipment_list, project_info
            )
            
            if success:
                QMessageBox.information(
                    self.parent, "导出成功",
                    f"已成功导出 {len(equipment_list)} 个设备 ({export_scope})\n文件已保存到:\n{output_path}"
                )
                
                # 询问是否打开文件
                reply = QMessageBox.question(
                    self.parent, "打开文件",
                    "是否立即打开导出的文件？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Yes:
                    self.parent.open_file(output_path)
            else:
                QMessageBox.warning(self.parent, "导出失败", "模板填充失败，请检查模板格式")
                
        except Exception as e:
            QMessageBox.critical(self.parent, "错误", 
                               f"导出过程中发生错误:\n{str(e)}")
    
    def _ask_export_scope(self, selected_count):
        """询问导出范围"""
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("选择导出范围")
        dialog.setMinimumWidth(300)
        
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(f"当前选中 {selected_count} 个设备"))
        layout.addWidget(QLabel("请选择要导出的范围:"))
        
        btn_layout = QVBoxLayout()
        
        selected_btn = QPushButton(f"导出选中设备 ({selected_count}个)")
        selected_btn.clicked.connect(lambda: self._set_scope_result(dialog, "selected"))
        btn_layout.addWidget(selected_btn)
        
        all_btn = QPushButton("导出全部设备")
        all_btn.clicked.connect(lambda: self._set_scope_result(dialog, "all"))
        btn_layout.addWidget(all_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(lambda: self._set_scope_result(dialog, "cancel"))
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.scope_result = None
        dialog.exec()
        
        return getattr(dialog, 'scope_result', 'cancel')
    
    def _set_scope_result(self, dialog, result):
        """设置导出范围结果"""
        dialog.scope_result = result
        dialog.accept()
    
    def _get_project_info_from_dialog(self):
        """获取项目信息"""
        from .equipment_dialogs import ProjectInfoDialog
        dialog = ProjectInfoDialog(self.parent)
        if dialog.exec() == QDialog.Accepted:
            return dialog.get_project_info()
        return None


class TemplateImportPreviewDialog(QDialog):
    """模板导入预览对话框"""
    def __init__(self, project_info, equipment_list, parent=None):
        super().__init__(parent)
        self.project_info = project_info
        self.equipment_list = equipment_list
        self.import_options = {
            'skip_existing': True,
            'update_existing': False,
            'import_all': True
        }
        
        self.setWindowTitle("模板导入预览")
        self.setMinimumSize(800, 600)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 项目信息区域
        if self.project_info:
            project_group = QGroupBox("项目信息")
            project_layout = QFormLayout(project_group)
            
            if 'project_name' in self.project_info:
                project_layout.addRow("项目名称:", QLabel(self.project_info['project_name']))
            
            if 'sub_item_name' in self.project_info:
                project_layout.addRow("子项名称:", QLabel(self.project_info['sub_item_name']))
            
            if 'doc_no' in self.project_info:
                project_layout.addRow("文件编号:", QLabel(self.project_info['doc_no']))
            
            if 'speciality' in self.project_info:
                project_layout.addRow("专业:", QLabel(self.project_info['speciality']))
            
            if 'phase' in self.project_info:
                project_layout.addRow("阶段:", QLabel(self.project_info['phase']))
            
            layout.addWidget(project_group)
        
        # 设备列表预览
        equipment_group = QGroupBox(f"设备列表 ({len(self.equipment_list)} 个设备)")
        equipment_layout = QVBoxLayout(equipment_group)
        
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(5)
        self.equipment_table.setHorizontalHeaderLabels([
            "序号", "设备位号", "设备名称", "规格", "唯一编码"
        ])
        
        self.equipment_table.setRowCount(len(self.equipment_list))
        for i, equipment in enumerate(self.equipment_list):
            self.equipment_table.setItem(i, 0, QTableWidgetItem(str(i+1)))
            self.equipment_table.setItem(i, 1, QTableWidgetItem(equipment.get('equipment_id', '')))
            self.equipment_table.setItem(i, 2, QTableWidgetItem(equipment.get('name', '')))
            self.equipment_table.setItem(i, 3, QTableWidgetItem(equipment.get('specification', '')[:50] + '...' 
                                                            if len(equipment.get('specification', '')) > 50 
                                                            else equipment.get('specification', '')))
            self.equipment_table.setItem(i, 4, QTableWidgetItem(equipment.get('unique_code', '')))
        
        self.equipment_table.horizontalHeader().setStretchLastSection(True)
        equipment_layout.addWidget(self.equipment_table)
        
        layout.addWidget(equipment_group)
        
        # 导入选项
        options_group = QGroupBox("导入选项")
        options_layout = QVBoxLayout(options_group)
        
        self.skip_existing_check = QCheckBox("跳过已存在的设备（根据设备位号）")
        self.skip_existing_check.setChecked(True)
        options_layout.addWidget(self.skip_existing_check)
        
        self.update_existing_check = QCheckBox("更新已存在的设备")
        self.update_existing_check.setChecked(False)
        options_layout.addWidget(self.update_existing_check)
        
        self.import_all_check = QCheckBox("导入所有设备")
        self.import_all_check.setChecked(True)
        options_layout.addWidget(self.import_all_check)
        
        layout.addWidget(options_group)
        
        # 按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
    
    def get_import_options(self):
        """获取导入选项"""
        self.import_options.update({
            'skip_existing': self.skip_existing_check.isChecked(),
            'update_existing': self.update_existing_check.isChecked(),
            'import_all': self.import_all_check.isChecked()
        })
        return self.import_options
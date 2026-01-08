# modules/process_design/tabs/equipment_templates.py
import os
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QPoint, QRect
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QComboBox, QPushButton, QDialogButtonBox, QFormLayout,
    QGroupBox, QCheckBox, QTableWidget, QTableWidgetItem,
    QTextEdit, QGroupBox, QScrollArea, QWidget, QToolTip,
    QTableWidget, QTableWidgetItem
)

from .equipment_list_tab import EquipmentIDGenerator


class EquipmentTemplateCreator:
    """设备清单模板创建器 - 针对实际Excel文件格式"""
    def __init__(self):
        self.templates_dir = "templates"
        os.makedirs(self.templates_dir, exist_ok=True)
        
        self.template_definitions = {
            "ACME模板": {
                "description": "包含完整信息的ACME设备清单模板",
                "data_start_row": 10,
                "columns": [
                    ("Item\n序号", "item", 7, "text"),
                    ("Tag num.\n设备位号", "equipment_id", 22, "text"),
                    ("Description", "description_en", 45, "text"),
                    ("设备名称", "name", 28, "text"),
                    ("P&ID DWG. NO.\nP&ID图号", "pid_dwg_no", 35, "text"),
                    ("Technical specifications\n设备技术规格", "specification", 60, "text"),
                    ("QTY.\n数量", "quantity", 10, "integer"),
                    ("Unit price\n单价\nCNY", "unit_price", 10, "currency"),
                    ("Total price\n总价\nCNY", "total_price", 10, "currency_formula"),
                    ("Design tem.\n设计温度\n℃", "design_temperature", 10, "number"),
                    ("Design pressure\n设计压力\nMPa·G", "design_pressure", 10, "number"),
                    ("Operating tem.\n操作温度\n℃", "operating_temperature", 10, "number"),
                    ("Operating pressure\n操作压力\nMPa·G", "operating_pressure", 10, "number"),
                    ("Unit power\n单机功率\nkW", "single_power", 10, "number"),
                    ("Run power\n运行功率\nkW", "operating_power", 10, "number"),
                    ("Installed power\n装机功率\nkW", "total_power", 10, "number"),
                    ("Material\n材质", "material", 10, "text"),
                    ("Insulation\n保温", "insulation", 10, "text"),
                    ("Dry weight\n单机重量\nt", "weight_estimate", 10, "number"),
                    ("Dynamic\n荷载系数", "dynamic", 10, "text"),
                    ("Remark\n备注", "notes", 20, "text"),
                    ("唯一编码", "unique_code", 30, "text")
                ]
            }
        }
    
    def create_template(self, template_type, file_path):
        """创建模板文件"""
        if template_type not in self.template_definitions:
            template_type = "ACME模板"
        
        template_def = self.template_definitions[template_type]
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "设备清单"
            
            self._create_actual_template(ws, template_def)
            
            wb.save(file_path)
            return file_path
            
        except Exception as e:
            raise
    
    def _create_actual_template(self, ws, template_def):
        """创建实际文件格式的模板"""
        
        # 第1行：项目名称和子项名称
        ws['A1'] = "项目名称"
        ws['D1'] = "子项名称"
        ws['Q1'] = "Doc. No. :"
        ws['Q2'] = "B-MD001"
        
        # 第3行：公司logo和名称、专业、设备清单标题
        ws['A3'] = "公司logo和名称"
        ws['D3'] = "SPECIALITY"
        ws['E3'] = "Mechnical"
        ws['F3'] = "EQUIPMENT LIST"
        ws['Q3'] = "REV."
        ws['S3'] = "A"
        ws['U3'] = "D/M/Y"
        
        # 第4行：阶段和签名信息
        ws['D4'] = "PHASE"
        ws['E4'] = "Basic"
        ws['Q4'] = "Prepared"
        ws['U4'] = "Approved"
        
        # 第5-8行：设备分类
        equipment_classes = [
            (5, "Standard equipment\n标准设备", "Heat exchanger\n换热器"),
            (6, "Complete sets & equipments\n成套设备", "Pump\n泵类"),
            (7, "Non-standard equipments\n非标设备", "Agitator\n搅拌"),
            (8, "Container class\n容器类", "")
        ]
        
        for row, class_cn, sub_class in equipment_classes:
            ws[f'A{row}'] = class_cn
            if sub_class:
                ws[f'D{row}'] = sub_class
        
        # 第9行：主要表头
        headers_row9 = {
            'A9': "Item\n序号",
            'B9': "Tag num.\n设备位号",
            'C9': "Description",
            'D9': "设备名称",
            'E9': "P&ID DWG. NO.\nP&ID图号",
            'F9': "Technical specifications\n设备技术规格",
            'G9': "QTY.\n数量",
            'H9': "Unit price\n单价\nCNY",
            'I9': "Total price\n总价\nCNY",
            'J9': "Design tem.\n设计温度\n℃",
            'K9': "Design pressure\n设计压力\nMPa·G",
            'L9': "Operating tem.\n操作温度\n℃",
            'M9': "Operating pressure\n操作压力\nMPa·G",
            'N9': "Unit power\n单机功率\nkW",
            'O9': "Run power\n运行功率\nkW",
            'P9': "Installed power\n装机功率\nkW",
            'Q9': "Material\n材质",
            'R9': "Insulation\n保温",
            'S9': "Dry weight\n单机重量\nt",
            'T9': "Operating weight\n操作重量\nt",
            'U9': "Total weight\总重量\nt",
            'V9': "Dynamic\n荷载系数",
            'W9': "Remark\n备注",
            'X9': "唯一编码"
        }
        
        for cell, value in headers_row9.items():
            ws[cell] = value
        
        # 合并单元格
        ws.merge_cells('A1:C2')  # 项目名称
        ws.merge_cells('D1:I2')  # 子项名称
        ws.merge_cells('Q1:W1')  # Doc. No.
        ws.merge_cells('Q2:W2')  # B-MD001
        ws.merge_cells('A3:C4')  # 公司logo和名称
        ws.merge_cells('F3:I4')  # EQUIPMENT LIST
        ws.merge_cells('Q3:R3')  # REV.
        ws.merge_cells('S3:T3')  # A
        ws.merge_cells('V3:W3')
        ws.merge_cells('Q4:R4')  # Prepared
        ws.merge_cells('S4:T4')
        ws.merge_cells('V4:W4')
        ws.merge_cells('A5:B5')
        ws.merge_cells('A6:B6')
        ws.merge_cells('A7:B7')
        ws.merge_cells('A8:B8')
        
        # 设置样式
        header_font = Font(bold=True, size=11)
        title_font = Font(name="微软雅黑", size=16, bold=True)
        
        # 设置标题样式
        ws['G3'].font = title_font
        ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置第9行表头样式
        for cell in headers_row9.keys():
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[cell].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 设置列宽
        column_widths = {
            'A': 7,    # Item
            'B': 22,   # Tag num.
            'C': 45,   # Description
            'D': 28,   # 设备名称
            'E': 35,   # P&ID DWG. NO.
            'F': 60,   # Technical specifications
            'G': 10,   # QTY.
            'H': 10,   # 单价
            'I': 10,   # 总价
            'J': 10,   # Design tem.
            'K': 10,   # Design pressure
            'L': 10,   # Operating tem.
            'M': 10,   # Operating pressure
            'N': 10,   # Estimated power(kW)
            'O': 10,   # Unit
            'P': 10,   # Run
            'Q': 10,   # Installed
            'R': 10,   # Material
            'S': 10,   # Insulation
            'T': 10,   # Weight estimate(t)
            'U': 10,   # Single
            'V': 10,   # Filling media
            'W': 20,   # Total, Dynamic, Remark
            'X': 30,
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 创建数据行边框
        data_start_row = template_def["data_start_row"]
        for i in range(data_start_row, data_start_row + 20):  # 创建20个数据行
            for col in range(1, 25):  # A到X列
                cell = ws.cell(row=i, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置数字列的对齐方式
                if col in [8, 9, 10, 11, 12, 13, 14, 15, 16, 21]:  # 数量、单价、总价等数值列
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if col in [9, 10]:  # 单价和总价
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def create_template_config(self, template_type):
        """创建模板配置文件"""
        config = {
            "template_type": template_type,
            "created_date": datetime.now().isoformat(),
            "version": "1.0",
            "description": self.template_definitions.get(template_type, {}).get("description", ""),
            "data_start_row": self.template_definitions.get(template_type, {}).get("data_start_row", 10),
            "columns": []
        }
        
        if template_type in self.template_definitions:
            for col_name, field, width, style in self.template_definitions[template_type]["columns"]:
                config["columns"].append({
                    "display_name": col_name,
                    "field_name": field,
                    "width": width,
                    "style": style
                })
        
        config_path = os.path.join(self.templates_dir, f"{template_type}_config.json")
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        return config_path


class EquipmentTemplateFiller:
    """设备清单模板填充器 - 针对实际Excel文件格式"""
    
    def fill_template(self, template_path, output_path, equipment_list, project_info):
        """
        填充模板文件
        
        参数:
            template_path: 模板文件路径
            output_path: 输出文件路径
            equipment_list: 设备列表
            project_info: 项目信息字典
        """
        try:
            # 加载模板工作簿
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 替换项目信息
            self._replace_project_info(ws, project_info)
            
            # 解析表头，获取列映射
            header_mapping = self._parse_actual_header(ws)
            
            # 确定数据开始行（通常是表头行的下一行）
            data_start_row = 10  # 根据ACME模板，数据从第10行开始
            
            # 填充设备数据
            self._fill_actual_equipment_data(ws, data_start_row, header_mapping, equipment_list)
            
            # 更新公式（如合计等）
            self._update_actual_formulas(ws, data_start_row, len(equipment_list))
            
            # 保存工作簿
            wb.save(output_path)
            
            return True
            
        except Exception as e:
            import traceback
            print(f"填充模板时发生错误: {e}")
            traceback.print_exc()
            return False
    
    def _replace_project_info(self, ws, project_info):
        """替换项目信息占位符 - 针对实际Excel文件格式"""
        
        # 第1行：项目名称和子项名称
        if project_info.get('project_name'):
            ws['A1'] = f"{project_info['project_name']}"
        
        if project_info.get('sub_item_name'):
            ws['D1'] = f"{project_info['sub_item_name']}"
        
        # Doc. No.
        if project_info.get('doc_no'):
            ws['Q2'] = project_info['doc_no']
        
        # 第3行：公司logo和名称、专业、设备清单标题
        # 公司logo和名称保持原样
        if project_info.get('speciality'):
            ws['E3'] = project_info['speciality']
        
        # 版次和日期
        if project_info.get('revision'):
            ws['S3'] = project_info['revision']
        
        if project_info.get('date'):
            ws['V3'] = project_info['date']
        
        # 第4行：阶段和签名信息
        if project_info.get('phase'):
            ws['E4'] = project_info['phase']
        
        # 设置单元格样式
        # 可以在这里添加样式设置代码
    
    def _parse_actual_header(self, ws):
        """解析实际文件格式的表头（单行表头）"""
        header_mapping = {}
        
        # 第9行是表头行
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=9, column=col).value
            
            if cell_value:
                header_text = str(cell_value)
                # 根据表头文本映射到字段名
                field_name = self._map_actual_header(header_text)
                if field_name:
                    header_mapping[field_name] = col
                    print(f"表头映射: 列{col} '{header_text}' -> {field_name}")
        
        return header_mapping
    
    def _map_actual_header(self, header_text):
        """映射实际表头到字段名"""
        mapping = {
            "Item\n序号": "item",
            "Tag num.\n设备位号": "equipment_id",
            "Description": "description_en",
            "设备名称": "name",
            "P&ID DWG. NO.\nP&ID图号": "pid_dwg_no",
            "Technical specifications\n设备技术规格": "specification",
            "QTY.\n数量": "quantity",
            "Unit price\n单价\nCNY": "unit_price",
            "Total price\n总价\nCNY": "total_price",
            "Design tem.\n设计温度\n℃": "design_temperature",
            "Design pressure\n设计压力\nMPa·G": "design_pressure",
            "Operating tem.\n操作温度\n℃": "operating_temperature",
            "Operating pressure\n操作压力\nMPa·G": "operating_pressure",
            "Unit power\n单机功率\nkW": "single_power",
            "Run power\n运行功率\nkW": "operating_power",
            "Installed power\n装机功率\nkW": "total_power",
            "Material\n材质": "material",
            "Insulation\n保温": "insulation",
            "Dry weight\n单机重量\nt": "weight_estimate",
            "Dynamic\n荷载系数": "dynamic",
            "Remark\n备注": "notes",
            "唯一编码": "unique_code"
        }
        
        for header, field in mapping.items():
            if header == header_text.strip():
                return field
        
        for header, field in mapping.items():
            # 移除括号和特殊字符进行比较
            clean_header = header.replace("(", "").replace(")", "").replace("·", "").strip()
            clean_text = header_text.replace("(", "").replace(")", "").replace("·", "").strip()
            
            if clean_header == clean_text or clean_header in clean_text or clean_text in clean_header:
                return field
        
        return None
    
    def _fill_actual_equipment_data(self, ws, start_row, header_mapping, equipment_list):
        """填充实际格式的设备数据"""
        
        for i, equipment in enumerate(equipment_list, start=0):
            row = start_row + i
            
            # 填充Item序号
            if "item" in header_mapping:
                ws.cell(row=row, column=header_mapping["item"], value=i+1)
            
            # 填充设备数据
            for field_name, col in header_mapping.items():
                if field_name == "item":
                    continue  # 已经处理
                
                try:
                    if field_name == "equipment_id":
                        value = getattr(equipment, 'equipment_id', '')
                        if not value and hasattr(equipment, 'name'):
                            value = f"EQ-{equipment.name.replace(' ', '_')}"
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "description_en":
                        value = getattr(equipment, 'description_en', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "name":
                        value = getattr(equipment, 'name', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "pid_dwg_no":
                        value = getattr(equipment, 'pid_dwg_no', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "specification":
                        value = getattr(equipment, 'specification', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "quantity":
                        value = getattr(equipment, 'quantity', 1)
                        ws.cell(row=row, column=col, value=value)
                        ws.cell(row=row, column=col).number_format = '0'
                    
                    elif field_name == "unit_price":
                        value = getattr(equipment, 'unit_price', 0)
                        ws.cell(row=row, column=col, value=value)
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                    
                    elif field_name == "total_price":
                        value = getattr(equipment, 'total_price', 0)
                        ws.cell(row=row, column=col, value=value)
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                    
                    elif field_name == "design_temperature":
                        value = getattr(equipment, 'design_temperature', '')
                        if value == "NT":
                            value = "NT"
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "design_pressure":
                        value = getattr(equipment, 'design_pressure', '')
                        if value == "NP":
                            value = "NP"
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "operating_temperature":
                        value = getattr(equipment, 'operating_temperature', '')
                        if value == "NT":
                            value = "NT"
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "operating_pressure":
                        value = getattr(equipment, 'operating_pressure', '')
                        if value == "NP":
                            value = "NP"
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "single_power":
                        value = getattr(equipment, 'single_power', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "operating_power":
                        value = getattr(equipment, 'operating_power', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "total_power":
                        value = getattr(equipment, 'total_power', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "material":
                        value = getattr(equipment, 'material', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "insulation":
                        value = getattr(equipment, 'insulation', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "weight_estimate":
                        value = getattr(equipment, 'weight_estimate', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "dynamic":
                        value = getattr(equipment, 'dynamic', '')
                        ws.cell(row=row, column=col, value=value)
                    
                    elif field_name == "notes":
                        value = getattr(equipment, 'notes', '')
                        ws.cell(row=row, column=col, value=value)
                        
                    elif field_name == "unique_code":
                        value = getattr(equipment, 'unique_code', '')
                        ws.cell(row=row, column=col, value=value)
                    
                except Exception as e:
                    print(f"填充字段 {field_name} 时出错: {e}")
    
    def _update_actual_formulas(self, ws, start_row, data_count):
        """更新实际文件格式的公式"""
        if data_count == 0:
            return
        
        end_row = start_row + data_count - 1
        
        # 在最后一行后添加汇总行
        summary_row = end_row + 2
        ws.cell(row=summary_row, column=1, value="合计:")
        
        # 计算总价汇总
        # 假设总价在J列（第10列）
        total_col = 10  # J列
        total_cell = ws.cell(row=summary_row, column=total_col)
        total_cell.value = f"=SUM(J{start_row}:J{end_row})"
        total_cell.number_format = '#,##0.00'


class TemplateImportPreviewDialog(QDialog):
    """模板导入预览对话框"""
    def __init__(self, project_info, equipment_list, parent=None):
        super().__init__(parent)
        self.project_info = project_info
        self.equipment_list = equipment_list
        self.import_options = {
            'skip_existing': True,
            'update_existing': False,
            'import_all': True
        }
        
        self.setWindowTitle("模板导入预览")
        self.setMinimumSize(800, 600)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 项目信息区域
        if self.project_info:
            project_group = QGroupBox("项目信息")
            project_layout = QFormLayout(project_group)
            
            if 'project_name' in self.project_info:
                project_layout.addRow("项目名称:", QLabel(self.project_info['project_name']))
            
            if 'sub_item_name' in self.project_info:
                project_layout.addRow("子项名称:", QLabel(self.project_info['sub_item_name']))
            
            if 'doc_no' in self.project_info:
                project_layout.addRow("文件编号:", QLabel(self.project_info['doc_no']))
            
            if 'speciality' in self.project_info:
                project_layout.addRow("专业:", QLabel(self.project_info['speciality']))
            
            if 'phase' in self.project_info:
                project_layout.addRow("阶段:", QLabel(self.project_info['phase']))
            
            layout.addWidget(project_group)
        
        # 设备列表预览
        equipment_group = QGroupBox(f"设备列表 ({len(self.equipment_list)} 个设备)")
        equipment_layout = QVBoxLayout(equipment_group)
        
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(5)
        self.equipment_table.setHorizontalHeaderLabels([
            "序号", "设备位号", "设备名称", "规格", "唯一编码"
        ])
        
        self.equipment_table.setRowCount(len(self.equipment_list))
        for i, equipment in enumerate(self.equipment_list):
            self.equipment_table.setItem(i, 0, QTableWidgetItem(str(i+1)))
            self.equipment_table.setItem(i, 1, QTableWidgetItem(equipment.get('equipment_id', '')))
            self.equipment_table.setItem(i, 2, QTableWidgetItem(equipment.get('name', '')))
            self.equipment_table.setItem(i, 3, QTableWidgetItem(equipment.get('specification', '')[:50] + '...' 
                                                            if len(equipment.get('specification', '')) > 50 
                                                            else equipment.get('specification', '')))
            self.equipment_table.setItem(i, 4, QTableWidgetItem(equipment.get('unique_code', '')))
        
        self.equipment_table.horizontalHeader().setStretchLastSection(True)
        equipment_layout.addWidget(self.equipment_table)
        
        layout.addWidget(equipment_group)
        
        # 导入选项
        options_group = QGroupBox("导入选项")
        options_layout = QVBoxLayout(options_group)
        
        self.skip_existing_check = QCheckBox("跳过已存在的设备（根据设备位号）")
        self.skip_existing_check.setChecked(True)
        options_layout.addWidget(self.skip_existing_check)
        
        self.update_existing_check = QCheckBox("更新已存在的设备")
        self.update_existing_check.setChecked(False)
        options_layout.addWidget(self.update_existing_check)
        
        self.import_all_check = QCheckBox("导入所有设备")
        self.import_all_check.setChecked(True)
        options_layout.addWidget(self.import_all_check)
        
        layout.addWidget(options_group)
        
        # 按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
    
    def get_import_options(self):
        """获取导入选项"""
        self.import_options.update({
            'skip_existing': self.skip_existing_check.isChecked(),
            'update_existing': self.update_existing_check.isChecked(),
            'import_all': self.import_all_check.isChecked()
        })
        return self.import_options


class ProjectInfoDialog(QDialog):
    """项目信息对话框 - 针对实际Excel文件格式"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("填写项目信息")
        self.setMinimumWidth(500)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout(basic_group)
        
        self.project_name_input = QLineEdit()
        self.project_name_input.setText("某化工项目")
        self.project_name_input.setPlaceholderText("项目名称")
        
        self.sub_item_name_input = QLineEdit()
        self.sub_item_name_input.setText(f"子项-{datetime.now().strftime('%Y%m')}-001")
        self.sub_item_name_input.setPlaceholderText("子项名称")
        
        self.doc_no_input = QLineEdit()
        self.doc_no_input.setText("B-MD001")
        self.doc_no_input.setPlaceholderText("文件编号")
        
        basic_layout.addRow("项目名称:", self.project_name_input)
        basic_layout.addRow("子项名称:", self.sub_item_name_input)
        basic_layout.addRow("文件编号(Doc. No.):", self.doc_no_input)
        
        layout.addWidget(basic_group)
        
        design_group = QGroupBox("设计信息")
        design_layout = QFormLayout(design_group)
        
        self.speciality_combo = QComboBox()
        self.speciality_combo.addItems(["Mechnical", "Process", "Electrical", "Instrument", "Civil"])
        self.speciality_combo.setCurrentText("Mechnical")
        
        self.phase_combo = QComboBox()
        self.phase_combo.addItems(["Basic", "Detailed", "Construction", "Feasibility"])
        self.phase_combo.setCurrentText("Basic")
        
        self.revision_input = QLineEdit()
        self.revision_input.setText("A")
        
        self.date_input = QLineEdit()
        self.date_input.setText(datetime.now().strftime("%Y/%m/%d"))
        
        design_layout.addRow("专业(SPECIALITY):", self.speciality_combo)
        design_layout.addRow("阶段(PHASE):", self.phase_combo)
        design_layout.addRow("版次(REV.):", self.revision_input)
        design_layout.addRow("日期(D/M/Y):", self.date_input)
        
        layout.addWidget(design_group)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_project_info(self):
        """获取项目信息 - 针对实际Excel文件格式"""
        return {
            # 基本信息
            "project_name": self.project_name_input.text().strip(),
            "sub_item_name": self.sub_item_name_input.text().strip(),
            "doc_no": self.doc_no_input.text().strip(),
            
            # 设计信息
            "speciality": self.speciality_combo.currentText(),
            "phase": self.phase_combo.currentText(),
            "revision": self.revision_input.text().strip(),
            "date": self.date_input.text().strip(),
            
            # 兼容旧版本的字段
            "project_code": self.sub_item_name_input.text().strip(),  # 子项名称作为项目编号
            "design_phase": self.phase_combo.currentText(),  # 阶段作为设计阶段
            "department": self.speciality_combo.currentText(),  # 专业作为部门
            "project_location": "",  # 实际模板中没有项目地点字段
        }


class TemplateTypeDialog(QDialog):
    """模板类型选择对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择模板类型")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("请选择要创建的模板类型:"))
        
        self.template_type_combo = QComboBox()
        self.template_type_combo.addItems([
            "ACME模板"
        ])
        
        layout.addWidget(self.template_type_combo)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_template_type(self):
        return self.template_type_combo.currentText()
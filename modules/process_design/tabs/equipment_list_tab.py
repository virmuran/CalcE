# modules/process_design/tabs/equipment_list_tab.py
import sys
import os
import subprocess

# 设置模块路径
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)  # 父目录 (process_design)
grandparent_dir = os.path.dirname(parent_dir)  # 祖父目录 (modules)

# 添加必要的路径到sys.path
paths_to_add = [
    current_dir,      # 当前目录
    parent_dir,       # 父目录（TofuApp\modules\process_design）
    grandparent_dir   # 祖父目录（TofuApp\modules）
]

for path in paths_to_add:
    if path not in sys.path:
        sys.path.insert(0, path)

from PySide6.QtCore import Qt, Signal, QTimer, QThread, QSize, QPoint, QRect  # 添加 QPoint 和 QRect
from PySide6.QtGui import QAction, QKeySequence, QClipboard
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QLabel, QHeaderView, QMessageBox, QDialog,
    QFormLayout, QDoubleSpinBox, QComboBox, QTextEdit, QGroupBox,
    QCheckBox, QFileDialog, QProgressDialog, QSplitter, QTabWidget,
    QMenu, QApplication, QFrame, QToolBar, QSizePolicy, QDialogButtonBox,
    QSpinBox, QScrollArea, QToolTip
)

# 导入其他库
import csv
import json
import pandas as pd
import re
from typing import List, Optional, Dict, Any
from datetime import datetime

# 导入openpyxl用于模板处理
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

# 尝试导入工艺设计相关模块
try:
    from ..process_design_manager import ProcessDesignManager
    from ..process_design_data import EquipmentItem
    print("✅ 成功导入 ProcessDesignManager 和 EquipmentItem")
except ImportError as e:
    print(f"❌ 导入失败: {e}")
    
    # 创建占位符类
    class ProcessDesignManager:
        def __init__(self, *args, **kwargs):
            pass
        def get_all_equipment(self):
            return []
        def get_equipment(self, equipment_id):
            return None
        def add_equipment(self, equipment):
            return False
        def update_equipment(self, equipment):
            return False
        def delete_equipment(self, equipment_id):
            return False
        def search_equipment(self, search_term):
            return []
        def advanced_search_equipment(self, criteria):
            return []
    
    class EquipmentItem:
        def __init__(self, **kwargs):
            for key, value in kwargs.items():
                setattr(self, key, value)

class EquipmentListTab(QWidget):
    """设备列表标签页"""
    
    equipment_selected = Signal(str)  # 设备选择信号
    equipment_list_updated = Signal()  # 设备列表更新信号
    import_progress = Signal(int)  # 导入进度信号
    
    def __init__(self, data_manager=None, parent=None):
        super().__init__(parent)
        self.data_manager = data_manager
        self.process_manager = None
        self.current_equipment = []  # 当前显示的设备列表
        self.selected_rows = set()  # 选中的行
        self.batch_mode = False  # 批量操作模式
        
        # 延迟初始化process_manager
        if data_manager:
            try:
                self.process_manager = ProcessDesignManager(data_manager)
                print("✅ 设备列表: ProcessDesignManager 初始化成功")
            except Exception as e:
                print(f"❌ 设备列表: ProcessDesignManager 初始化失败: {e}")
                self.process_manager = None
        
        self.setup_ui()
        self.load_equipment()
        self.setup_shortcuts()
    
    def setup_ui(self):
        """设置UI"""
        main_layout = QVBoxLayout(self)
        
        # 工具栏
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(16, 16))
        
        # 工具栏动作
        self.add_action = QAction("添加", self)
        self.add_action.triggered.connect(self.add_equipment)
        toolbar.addAction(self.add_action)
        
        self.edit_action = QAction("编辑", self)
        self.edit_action.triggered.connect(self.edit_equipment)
        toolbar.addAction(self.edit_action)
        
        self.delete_action = QAction("删除", self)
        self.delete_action.triggered.connect(self.delete_equipment)
        toolbar.addAction(self.delete_action)
        
        toolbar.addSeparator()
        
        self.import_action = QAction("导入", self)
        self.import_action.triggered.connect(self.import_equipment)
        toolbar.addAction(self.import_action)
        
        self.export_action = QAction("导出", self)
        self.export_action.triggered.connect(self.export_equipment)
        toolbar.addAction(self.export_action)
        
        toolbar.addSeparator()
        
        self.batch_toggle_action = QAction("批量模式", self)
        self.batch_toggle_action.setCheckable(True)
        self.batch_toggle_action.toggled.connect(self.toggle_batch_mode)
        toolbar.addAction(self.batch_toggle_action)
        
        toolbar.addSeparator()
        
        self.mapping_action = QAction("对照表", self)
        self.mapping_action.triggered.connect(self.manage_name_mapping)
        toolbar.addAction(self.mapping_action)
        
        # === 新增：模板相关按钮 ===
        toolbar.addSeparator()
        
        # 模板创建按钮
        self.template_create_action = QAction("创建模板", self)
        self.template_create_action.triggered.connect(self.create_template)
        toolbar.addAction(self.template_create_action)
        
        # 模板导出按钮
        self.template_export_action = QAction("模板导出", self)
        self.template_export_action.triggered.connect(self.export_equipment_with_template)
        toolbar.addAction(self.template_export_action)
        
        main_layout.addWidget(toolbar)
        
        # 搜索和过滤区域
        filter_frame = QFrame()
        filter_frame.setFrameStyle(QFrame.StyledPanel)
        filter_layout = QHBoxLayout(filter_frame)
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_label = QLabel("搜索:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("设备名称、型号或编号...")
        self.search_input.textChanged.connect(self.on_search_changed)
        self.search_input.returnPressed.connect(self.perform_search)
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        
        # 高级搜索按钮
        self.advanced_search_btn = QPushButton("高级搜索")
        self.advanced_search_btn.clicked.connect(self.open_advanced_search)
        search_layout.addWidget(self.advanced_search_btn)
        
        filter_layout.addLayout(search_layout)
        
        # 过滤器
        filter_layout.addStretch()
        
        # 设备类型过滤器
        self.type_filter = QComboBox()
        self.type_filter.addItem("所有类型")
        self.type_filter.addItems(["反应器", "分离器", "换热器", "泵", "压缩机", "储罐", "阀门", "管道", "其他"])
        self.type_filter.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("设备类型:"))
        filter_layout.addWidget(self.type_filter)
        
        # 状态过滤器
        self.status_filter = QComboBox()
        self.status_filter.addItem("所有状态")
        self.status_filter.addItems(["运行中", "停机", "维修中", "备用"])
        self.status_filter.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("状态:"))
        filter_layout.addWidget(self.status_filter)
        
        main_layout.addWidget(filter_frame)
        
        # 分割器：左侧表格，右侧详情
        splitter = QSplitter(Qt.Horizontal)
        
        # 左侧：设备表格
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        
        # 表格上方信息栏
        info_layout = QHBoxLayout()
        self.info_label = QLabel("总计: 0 个设备")
        info_layout.addWidget(self.info_label)
        info_layout.addStretch()
        
        self.selected_label = QLabel("已选择: 0 个")
        info_layout.addWidget(self.selected_label)
        table_layout.addLayout(info_layout)
        
        # 设备表格
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(13)
        self.equipment_table.setHorizontalHeaderLabels([
            "",  # 选择框
            "设备ID", "设备名称", "设备类型", "型号", "规格", 
            "制造商", "安装位置", "状态", "投用日期", 
            "设计压力", "设计温度", "备注"
        ])
        
        # 设置表头
        header = self.equipment_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # 选择列固定宽度
        header.resizeSection(0, 30)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # ID列自适应
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # 名称列拉伸
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 类型
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 型号
        
        # 启用排序
        self.equipment_table.setSortingEnabled(True)
        
        # 设置选择模式
        self.equipment_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.equipment_table.setSelectionMode(QTableWidget.ExtendedSelection)
        
        # 连接信号
        self.equipment_table.itemDoubleClicked.connect(self.on_equipment_double_clicked)
        self.equipment_table.itemSelectionChanged.connect(self.on_selection_changed)
        
        table_layout.addWidget(self.equipment_table)
        
        splitter.addWidget(table_widget)
        
        # 右侧：设备详情
        detail_widget = QWidget()
        detail_layout = QVBoxLayout(detail_widget)
        
        # 详情标签
        detail_label = QLabel("设备详情")
        detail_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        detail_layout.addWidget(detail_label)
        
        # 详情显示区域
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setMaximumHeight(300)
        detail_layout.addWidget(self.detail_text)
        
        # 技术参数标签
        property_label = QLabel("技术参数")
        property_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        detail_layout.addWidget(property_label)
        
        # 技术参数表格
        self.property_table = QTableWidget()
        self.property_table.setColumnCount(2)
        self.property_table.setHorizontalHeaderLabels(["参数", "值"])
        self.property_table.horizontalHeader().setStretchLastSection(True)
        self.property_table.setEditTriggers(QTableWidget.NoEditTriggers)
        detail_layout.addWidget(self.property_table)
        
        splitter.addWidget(detail_widget)
        splitter.setSizes([700, 300])  # 设置初始大小比例
        
        main_layout.addWidget(splitter)
        
        # 批量操作按钮（初始隐藏）
        self.batch_panel = QFrame()
        self.batch_panel.setFrameStyle(QFrame.StyledPanel)
        self.batch_panel.setVisible(False)
        batch_layout = QHBoxLayout(self.batch_panel)
        
        self.batch_edit_btn = QPushButton("批量编辑")
        self.batch_edit_btn.clicked.connect(self.batch_edit_equipment)
        batch_layout.addWidget(self.batch_edit_btn)
        
        self.batch_export_btn = QPushButton("批量导出")
        self.batch_export_btn.clicked.connect(self.batch_export_equipment)
        batch_layout.addWidget(self.batch_export_btn)
        
        self.batch_delete_btn = QPushButton("批量删除")
        self.batch_delete_btn.clicked.connect(self.batch_delete_equipment)
        batch_layout.addWidget(self.batch_delete_btn)
        
        batch_layout.addStretch()
        
        self.clear_batch_btn = QPushButton("清除选择")
        self.clear_batch_btn.clicked.connect(self.clear_batch_selection)
        batch_layout.addWidget(self.clear_batch_btn)
        
        main_layout.addWidget(self.batch_panel)
        
        # 状态栏
        self.status_bar = QLabel()
        self.status_bar.setFrameStyle(QFrame.StyledPanel | QFrame.Sunken)
        main_layout.addWidget(self.status_bar)
    
    def setup_shortcuts(self):
        """设置快捷键"""
        # 复制快捷键
        self.copy_action = QAction("复制", self)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.copy_action.triggered.connect(self.copy_selected)
        self.addAction(self.copy_action)
        
        # 刷新快捷键
        self.refresh_action = QAction("刷新", self)
        self.refresh_action.setShortcut(QKeySequence.Refresh)
        self.refresh_action.triggered.connect(self.load_equipment)
        self.addAction(self.refresh_action)
    
    # ==================== 模板相关方法 ====================
    
    def export_equipment_with_template(self):
        """使用模板导出设备清单"""
        try:
            # 选择模板文件
            template_path, _ = QFileDialog.getOpenFileName(
                self, "选择模板文件", "",
                "Excel模板文件 (*.xlsx);;所有文件 (*)"
            )
            
            if not template_path:
                return
            
            # 选择输出文件
            default_name = f"设备清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path, _ = QFileDialog.getSaveFileName(
                self, "保存设备清单", default_name,
                "Excel文件 (*.xlsx)"
            )
            
            if not output_path:
                return
            
            # 创建模板填充器
            template_filler = EquipmentTemplateFiller()
            
            # 获取要导出的设备
            equipment_list = self.get_equipment_to_export()
            
            if not equipment_list:
                QMessageBox.warning(self, "警告", "没有设备可导出")
                return
            
            # 获取项目信息
            project_info = self.get_project_info_from_dialog()
            
            if project_info is None:
                return
            
            # 填充模板
            success = template_filler.fill_template(
                template_path, output_path, equipment_list, project_info
            )
            
            if success:
                QMessageBox.information(
                    self, "导出成功",
                    f"已成功导出 {len(equipment_list)} 个设备\n文件已保存到:\n{output_path}"
                )
                
                # 询问是否打开文件
                reply = QMessageBox.question(
                    self, "打开文件",
                    "是否立即打开导出的文件？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Yes:
                    self.open_file(output_path)
            else:
                QMessageBox.warning(self, "导出失败", "模板填充失败，请检查模板格式")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出过程中发生错误:\n{str(e)}")
    
    def get_equipment_to_export(self):
        """获取要导出的设备列表"""
        if self.batch_mode:
            # 批量模式：获取选中的设备
            selected_ids = self.get_selected_equipment()
            equipment_list = []
            for equipment_id in selected_ids:
                equipment = self.process_manager.get_equipment(equipment_id)
                if equipment:
                    equipment_list.append(equipment)
            return equipment_list
        else:
            # 普通模式：获取所有设备
            return self.process_manager.get_all_equipment()
    
    def get_project_info_from_dialog(self):
        """从对话框获取项目信息"""
        dialog = ProjectInfoDialog(self)
        if dialog.exec() == QDialog.Accepted:
            return dialog.get_project_info()
        return None
    
    def open_file(self, file_path):
        """打开文件"""
        try:
            if sys.platform == "win32":
                os.startfile(file_path)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", file_path])
            else:  # linux
                subprocess.run(["xdg-open", file_path])
        except Exception as e:
            print(f"打开文件失败: {e}")
    
    def create_template(self):
        """创建设备清单模板"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存模板文件", "设备清单模板.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 创建模板生成器
            creator = EquipmentTemplateCreator()
            
            # 选择模板类型
            dialog = TemplateTypeDialog(self)
            if dialog.exec() == QDialog.Accepted:
                template_type = dialog.get_template_type()
                
                # 创建模板
                template_path = creator.create_template(template_type, file_path)
                
                # 创建配置文件
                config_path = creator.create_template_config(template_type)
                
                QMessageBox.information(
                    self, "模板创建成功",
                    f"模板已创建:\n{template_path}\n\n配置文件:\n{config_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"创建模板失败:\n{str(e)}")
    
    # ==================== 原有方法 ====================
    
    def load_equipment(self):
        """加载设备数据"""
        if not self.process_manager:
            self.status_bar.setText("错误: 数据管理器未初始化")
            return
        
        try:
            self.current_equipment = self.process_manager.get_all_equipment()
            self.populate_table(self.current_equipment)
            self.update_info_label()
            self.status_bar.setText(f"数据加载完成: {len(self.current_equipment)} 条记录")
        except Exception as e:
            self.status_bar.setText(f"加载失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"加载设备数据时发生错误:\n{str(e)}")
    
    def populate_table(self, equipment_list):
        """填充表格数据"""
        self.equipment_table.setRowCount(len(equipment_list))
        
        for i, equipment in enumerate(equipment_list):
            # 选择框
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox_item.setCheckState(Qt.Unchecked)
            self.equipment_table.setItem(i, 0, checkbox_item)
            
            # 设备数据
            self.equipment_table.setItem(i, 1, QTableWidgetItem(equipment.equipment_id))
            self.equipment_table.setItem(i, 2, QTableWidgetItem(equipment.name))
            self.equipment_table.setItem(i, 3, QTableWidgetItem(equipment.equipment_type if hasattr(equipment, 'equipment_type') else equipment.type))
            self.equipment_table.setItem(i, 4, QTableWidgetItem(equipment.model))
            self.equipment_table.setItem(i, 5, QTableWidgetItem(equipment.specification))
            self.equipment_table.setItem(i, 6, QTableWidgetItem(equipment.manufacturer))
            self.equipment_table.setItem(i, 7, QTableWidgetItem(equipment.location))
            self.equipment_table.setItem(i, 8, QTableWidgetItem(equipment.status))
            
            # 日期格式化
            if hasattr(equipment, 'commission_date') and equipment.commission_date:
                date_str = equipment.commission_date.strftime("%Y-%m-%d") if hasattr(equipment.commission_date, 'strftime') else str(equipment.commission_date)
            else:
                date_str = ""
            self.equipment_table.setItem(i, 9, QTableWidgetItem(date_str))
            
            # 设计参数 - 修复：处理字符串和数值类型
            if hasattr(equipment, 'design_pressure') and equipment.design_pressure:
                if isinstance(equipment.design_pressure, (int, float)):
                    pressure_str = f"{equipment.design_pressure:.2f}"
                else:
                    pressure_str = str(equipment.design_pressure)
            else:
                pressure_str = ""
            self.equipment_table.setItem(i, 10, QTableWidgetItem(pressure_str))
            
            if hasattr(equipment, 'design_temperature') and equipment.design_temperature:
                if isinstance(equipment.design_temperature, (int, float)):
                    temp_str = f"{equipment.design_temperature:.1f}"
                else:
                    temp_str = str(equipment.design_temperature)
            else:
                temp_str = ""
            self.equipment_table.setItem(i, 11, QTableWidgetItem(temp_str))
            
            self.equipment_table.setItem(i, 12, QTableWidgetItem(equipment.notes or ""))
        
        self.update_info_label()
    
    def update_info_label(self):
        """更新信息标签"""
        total = self.equipment_table.rowCount()
        selected = len([i for i in range(total) 
                       if self.equipment_table.item(i, 0).checkState() == Qt.Checked])
        self.info_label.setText(f"总计: {total} 个设备")
        self.selected_label.setText(f"已选择: {selected} 个")
    
    def on_search_changed(self):
        """搜索内容变化 - 防抖处理"""
        if hasattr(self, '_search_timer'):
            self._search_timer.stop()
        
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self.perform_search)
        self._search_timer.start(500)  # 500ms防抖
    
    def perform_search(self):
        """执行搜索"""
        search_term = self.search_input.text().strip()
        if not self.process_manager:
            return
        
        try:
            if not search_term:
                equipment_list = self.process_manager.get_all_equipment()
            else:
                equipment_list = self.process_manager.search_equipment(search_term)
            
            self.current_equipment = equipment_list
            self.apply_filters()  # 应用当前过滤器
            self.status_bar.setText(f"搜索到 {len(equipment_list)} 条记录")
        except Exception as e:
            self.status_bar.setText(f"搜索失败: {str(e)}")
    
    def apply_filters(self):
        """应用过滤器"""
        type_filter = self.type_filter.currentText()
        status_filter = self.status_filter.currentText()
        
        filtered_equipment = self.current_equipment.copy()
        
        # 应用设备类型过滤
        if type_filter != "所有类型":
            filtered_equipment = [e for e in filtered_equipment if e.equipment_type == type_filter]
        
        # 应用状态过滤
        if status_filter != "所有状态":
            filtered_equipment = [e for e in filtered_equipment if e.status == status_filter]
        
        self.populate_table(filtered_equipment)
    
    def open_advanced_search(self):
        """打开高级搜索对话框"""
        dialog = AdvancedSearchDialog(self)
        if dialog.exec() == QDialog.Accepted:
            criteria = dialog.get_search_criteria()
            self.perform_advanced_search(criteria)
    
    def perform_advanced_search(self, criteria):
        """执行高级搜索"""
        if not self.process_manager:
            return
        
        try:
            equipment_list = self.process_manager.advanced_search_equipment(criteria)
            self.current_equipment = equipment_list
            self.populate_table(equipment_list)
            self.status_bar.setText(f"高级搜索找到 {len(equipment_list)} 条记录")
        except Exception as e:
            self.status_bar.setText(f"高级搜索失败: {str(e)}")
            QMessageBox.warning(self, "搜索失败", str(e))
    
    def on_equipment_double_clicked(self, item):
        """设备双击事件"""
        if item.column() == 0:  # 点击选择框时不触发
            return
        
        row = item.row()
        equipment_id = self.equipment_table.item(row, 1).text()
        self.show_equipment_details(equipment_id)
        self.equipment_selected.emit(equipment_id)
    
    def show_equipment_details(self, equipment_id):
        """显示设备详情"""
        if not self.process_manager:
            return
        
        equipment = self.process_manager.get_equipment(equipment_id)
        if not equipment:
            return
        
        # 显示基本信息
        details = f"<h3>{equipment.name} ({equipment.equipment_id})</h3>"
        details += f"<b>设备类型:</b> {equipment.equipment_type if hasattr(equipment, 'equipment_type') else equipment.type}<br>"
        details += f"<b>型号:</b> {equipment.model}<br>"
        details += f"<b>规格:</b> {equipment.specification}<br>"
        details += f"<b>制造商:</b> {equipment.manufacturer}<br>"
        details += f"<b>安装位置:</b> {equipment.location}<br>"
        details += f"<b>状态:</b> {equipment.status}<br>"
        
        if hasattr(equipment, 'commission_date') and equipment.commission_date:
            details += f"<b>投用日期:</b> {equipment.commission_date.strftime('%Y-%m-%d') if hasattr(equipment.commission_date, 'strftime') else equipment.commission_date}<br>"
        
        if equipment.notes:
            details += f"<br><b>备注:</b><br>{equipment.notes}"
        
        self.detail_text.setHtml(details)
        
        # 显示技术参数表格
        # 处理设计压力和设计温度（可能是字符串"NP"、"NT"或数字）
        design_pressure_str = self._format_parameter(equipment.design_pressure, "MPa", 2) if hasattr(equipment, 'design_pressure') else "未知"
        design_temperature_str = self._format_parameter(equipment.design_temperature, "°C", 1) if hasattr(equipment, 'design_temperature') else "未知"
        operating_pressure_str = self._format_parameter(equipment.operating_pressure, "MPa", 2) if hasattr(equipment, 'operating_pressure') else "未知"
        operating_temperature_str = self._format_parameter(equipment.operating_temperature, "°C", 1) if hasattr(equipment, 'operating_temperature') else "未知"
        
        properties = [
            ("设计压力", design_pressure_str),
            ("设计温度", design_temperature_str),
            ("操作压力", operating_pressure_str),
            ("操作温度", operating_temperature_str),
            ("容积", f"{equipment.volume:.2f} m³" if hasattr(equipment, 'volume') and equipment.volume else "未知"),
            ("材质", equipment.material if hasattr(equipment, 'material') and equipment.material else "未知"),
            ("功率", f"{equipment.power:.1f} kW" if hasattr(equipment, 'power') and equipment.power else "未知"),
        ]
        
        self.property_table.setRowCount(len(properties))
        for i, (prop, value) in enumerate(properties):
            self.property_table.setItem(i, 0, QTableWidgetItem(prop))
            self.property_table.setItem(i, 1, QTableWidgetItem(value))
            
    def _format_parameter(self, value, unit, decimals):
        """格式化参数值，处理字符串和数值类型"""
        if value is None:
            return "未知"
        
        if isinstance(value, (int, float)):
            # 数值类型，格式化显示
            if decimals == 1:
                return f"{value:.1f} {unit}"
            elif decimals == 2:
                return f"{value:.2f} {unit}"
            else:
                return f"{value} {unit}"
        else:
            # 字符串类型（如"NP"、"NT"），直接返回字符串
            return f"{value} {unit}"
    
    def on_selection_changed(self):
        """选择变化事件"""
        selected_rows = self.equipment_table.selectionModel().selectedRows()
        if selected_rows:
            row = selected_rows[0].row()
            equipment_id = self.equipment_table.item(row, 1).text()
            self.show_equipment_details(equipment_id)
    
    def toggle_batch_mode(self, enabled):
        """切换批量操作模式"""
        self.batch_mode = enabled
        self.batch_panel.setVisible(enabled)
        
        if enabled:
            self.equipment_table.setSelectionMode(QTableWidget.NoSelection)
            for i in range(self.equipment_table.rowCount()):
                item = self.equipment_table.item(i, 0)
                item.setFlags(item.flags() | Qt.ItemIsEnabled)
        else:
            self.equipment_table.setSelectionMode(QTableWidget.ExtendedSelection)
            self.clear_batch_selection()
    
    def clear_batch_selection(self):
        """清除批量选择"""
        for i in range(self.equipment_table.rowCount()):
            self.equipment_table.item(i, 0).setCheckState(Qt.Unchecked)
        self.update_info_label()
    
    def get_selected_equipment(self):
        """获取选中的设备ID列表"""
        selected = []
        for i in range(self.equipment_table.rowCount()):
            if self.equipment_table.item(i, 0).checkState() == Qt.Checked:
                equipment_id = self.equipment_table.item(i, 1).text()
                selected.append(equipment_id)
        return selected
    
    def import_equipment(self):
        """导入设备数据"""
        file_path, selected_filter = QFileDialog.getOpenFileName(
            self, "导入设备数据", "", 
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;JSON文件 (*.json)"
        )
        
        if not file_path:
            return
        
        try:
            imported_count = 0
            
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # 读取Excel文件
                df = pd.read_excel(file_path, dtype=str)
                
                # 处理列名映射
                column_mapping = {
                    'Tag num.': 'equipment_id',
                    'Description': 'description_en',
                    '设备名称': 'name',
                    'Technical specifications': 'specification',
                    'Design tem.℃': 'design_temperature',
                    'Design pressure MPa·G': 'design_pressure',
                    'Operating tem.℃': 'operating_temperature',
                    'Operating pressure MPa·G': 'operating_pressure',
                    'Estimated power(kW)': 'estimated_power',
                    'Material': 'material',
                    'Insulation': 'insulation',
                    'Weight estimate(t)': 'weight_estimate',
                    'Remark': 'notes'
                }
                
                # 重命名列
                df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns}, inplace=True)
                
                for _, row in df.iterrows():
                    # 创建设备对象
                    equipment_data = {}
                    for col in df.columns:
                        if col in ['equipment_id', 'name', 'description_en']:
                            equipment_data[col] = str(row[col]) if pd.notna(row[col]) else ""
                        elif col in ['design_temperature', 'design_pressure', 'operating_temperature', 
                                    'operating_pressure', 'estimated_power', 'weight_estimate']:
                            try:
                                equipment_data[col] = float(row[col]) if pd.notna(row[col]) else None
                            except:
                                equipment_data[col] = None
                        else:
                            equipment_data[col] = str(row[col]) if pd.notna(row[col]) else ""
                    
                    # 设置默认值
                    if 'type' not in equipment_data:
                        equipment_data['type'] = '其他'
                    if 'status' not in equipment_data:
                        equipment_data['status'] = '运行中'
                    
                    # 添加或更新对照表
                    if equipment_data.get('name') and equipment_data.get('description_en'):
                        self.data_manager.add_equipment_name_mapping(
                            equipment_data['name'], 
                            equipment_data['description_en']
                        )
                    
                    # 创建设备对象并保存
                    from ..process_design_data import EquipmentItem
                    equipment = EquipmentItem(**equipment_data)
                    
                    if self.process_manager.add_equipment(equipment):
                        imported_count += 1
            
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8', dtype=str)
                # 处理逻辑与Excel类似
                column_mapping = {
                    'Tag num.': 'equipment_id',
                    'Description': 'description_en',
                    '设备名称': 'name',
                    'Technical specifications': 'specification',
                    'Design tem.℃': 'design_temperature',
                    'Design pressure MPa·G': 'design_pressure',
                    'Operating tem.℃': 'operating_temperature',
                    'Operating pressure MPa·G': 'operating_pressure',
                    'Estimated power(kW)': 'estimated_power',
                    'Material': 'material',
                    'Insulation': 'insulation',
                    'Weight estimate(t)': 'weight_estimate',
                    'Remark': 'notes'
                }
                
                df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns}, inplace=True)
                
                for _, row in df.iterrows():
                    equipment_data = {}
                    for col in df.columns:
                        if col in ['equipment_id', 'name', 'description_en']:
                            equipment_data[col] = str(row[col]) if pd.notna(row[col]) else ""
                        elif col in ['design_temperature', 'design_pressure', 'operating_temperature', 
                                    'operating_pressure', 'estimated_power', 'weight_estimate']:
                            try:
                                equipment_data[col] = float(row[col]) if pd.notna(row[col]) else None
                            except:
                                equipment_data[col] = None
                        else:
                            equipment_data[col] = str(row[col]) if pd.notna(row[col]) else ""
                    
                    if 'type' not in equipment_data:
                        equipment_data['type'] = '其他'
                    if 'status' not in equipment_data:
                        equipment_data['status'] = '运行中'
                    
                    if equipment_data.get('name') and equipment_data.get('description_en'):
                        self.data_manager.add_equipment_name_mapping(
                            equipment_data['name'], 
                            equipment_data['description_en']
                        )
                    
                    from ..process_design_data import EquipmentItem
                    equipment = EquipmentItem(**equipment_data)
                    
                    if self.process_manager.add_equipment(equipment):
                        imported_count += 1
            
            elif file_path.endswith('.json'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                for item in data:
                    from ..process_design_data import EquipmentItem
                    equipment = EquipmentItem.from_dict(item)
                    
                    # 添加对照表
                    if equipment.name and equipment.description_en:
                        self.data_manager.add_equipment_name_mapping(
                            equipment.name, 
                            equipment.description_en
                        )
                    
                    if self.process_manager.add_equipment(equipment):
                        imported_count += 1
            
            self.load_equipment()
            self.equipment_list_updated.emit()
            
            QMessageBox.information(self, "导入成功", f"成功导入 {imported_count} 个设备")
            
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入文件时发生错误:\n{str(e)}")
    
    def export_equipment(self):
        """导出设备数据（按照设备清单模板格式）"""
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, "导出设备清单", "设备清单.xlsx",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv);;JSON文件 (*.json)"
        )
        
        if not file_path:
            return
        
        try:
            # 获取所有设备
            equipment_list = self.process_manager.get_all_equipment()
            
            # 按照模板格式构建数据
            rows = []
            for i, equipment in enumerate(equipment_list, 1):
                # 获取英文描述
                description_en = ""
                if hasattr(equipment, 'description_en') and equipment.description_en:
                    description_en = equipment.description_en
                elif equipment.name:
                    # 尝试从对照表获取
                    description_en = self.data_manager.get_english_name(equipment.name)
                
                row = {
                    'Item': i,
                    'Tag num.': equipment.equipment_id,
                    'Description': description_en,
                    '设备名称': equipment.name,
                    'P&ID DWG. NO.': equipment.pid_dwg_no if hasattr(equipment, 'pid_dwg_no') else '',
                    'Technical specifications': equipment.specification,
                    'QTY.': equipment.quantity if hasattr(equipment, 'quantity') else 1,
                    '单价\nPrice': equipment.unit_price if hasattr(equipment, 'unit_price') else '',
                    '总价\nTotal': equipment.total_price if hasattr(equipment, 'total_price') else '',
                    'Design tem.<br>℃': equipment.design_temperature,
                    'Design pressure<br>MPa·G': equipment.design_pressure,
                    'Operating tem.<br>℃': equipment.operating_temperature if hasattr(equipment, 'operating_temperature') else '',
                    'Operating pressure<br>MPa·G': equipment.operating_pressure if hasattr(equipment, 'operating_pressure') else '',
                    'Estimated power(kW)': equipment.estimated_power if hasattr(equipment, 'estimated_power') else '',
                    'Material': equipment.material if hasattr(equipment, 'material') else '',
                    'Insulation': equipment.insulation if hasattr(equipment, 'insulation') else '',
                    'Weight estimate(t)': equipment.weight_estimate if hasattr(equipment, 'weight_estimate') else '',
                    'Dynamic': equipment.dynamic if hasattr(equipment, 'dynamic') else '',
                    'Remark': equipment.notes
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            
            # 设置列顺序（按照模板顺序）
            columns = [
                'Item', 'Tag num.', 'Description', '设备名称', 'P&ID DWG. NO.', 
                'Technical specifications', 'QTY.', '单价\nPrice', '总价\nTotal',
                'Design tem.<br>℃', 'Design pressure<br>MPa·G', 'Operating tem.<br>℃', 
                'Operating pressure<br>MPa·G', 'Estimated power(kW)', 
                'Material', 'Insulation', 'Weight estimate(t)', 
                'Dynamic', 'Remark'
            ]
            
            # 重新排列列顺序
            df = df.reindex(columns=[col for col in columns if col in df.columns])
            
            # 根据文件类型保存
            if file_path.endswith('.xlsx'):
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    
                    # 设置列宽（可选）
                    worksheet = writer.sheets['Sheet1']
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                        
            elif file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
            elif file_path.endswith('.json'):
                df.to_json(file_path, orient='records', force_ascii=False, indent=2)
            
            QMessageBox.information(self, "导出成功", f"成功导出 {len(equipment_list)} 个设备")
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出文件时发生错误:\n{str(e)}")
    
    def copy_selected(self):
        """复制选中内容到剪贴板"""
        selected_items = self.equipment_table.selectedItems()
        if not selected_items:
            return
        
        # 获取选中的行和列
        rows = sorted(set(item.row() for item in selected_items))
        cols = sorted(set(item.column() for item in selected_items))
        
        # 构建表格文本
        text = ""
        for row in rows:
            row_data = []
            for col in cols:
                if col == 0:  # 跳过选择框列
                    continue
                item = self.equipment_table.item(row, col)
                row_data.append(item.text() if item else "")
            text += "\t".join(row_data) + "\n"
        
        clipboard = QApplication.clipboard()
        clipboard.setText(text.strip())
        
        self.status_bar.setText(f"已复制 {len(rows)} 行数据")
    
    def batch_edit_equipment(self):
        """批量编辑设备"""
        selected_ids = self.get_selected_equipment()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要编辑的设备")
            return
        
        dialog = BatchEditDialog(selected_ids, self.process_manager, self)
        if dialog.exec() == QDialog.Accepted:
            self.load_equipment()
    
    def batch_export_equipment(self):
        """批量导出设备"""
        selected_ids = self.get_selected_equipment()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要导出的设备")
            return
        
        self.export_selected_equipment(selected_ids)
    
    def batch_delete_equipment(self):
        """批量删除设备"""
        selected_ids = self.get_selected_equipment()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要删除的设备")
            return
        
        reply = QMessageBox.question(
            self, "确认批量删除",
            f"确定要删除选中的 {len(selected_ids)} 个设备吗？\n此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success_count = 0
            for equipment_id in selected_ids:
                if self.process_manager.delete_equipment(equipment_id):
                    success_count += 1
            
            self.load_equipment()
            self.equipment_list_updated.emit()
            
            QMessageBox.information(
                self, "删除完成",
                f"已删除 {success_count} 个设备"
            )
    
    def export_selected_equipment(self, equipment_ids):
        """导出选中的设备"""
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, "导出选中设备", "selected_equipment.xlsx",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv);;JSON文件 (*.json)"
        )
        
        if not file_path:
            return
        
        try:
            equipment_list = []
            for equipment_id in equipment_ids:
                equipment = self.process_manager.get_equipment(equipment_id)
                if equipment:
                    equipment_list.append(equipment)
            
            # 按照模板格式构建数据
            rows = []
            for i, equipment in enumerate(equipment_list, 1):
                # 获取英文描述
                description_en = ""
                if hasattr(equipment, 'description_en') and equipment.description_en:
                    description_en = equipment.description_en
                elif equipment.name:
                    description_en = self.data_manager.get_english_name(equipment.name)
                
                row = {
                    'Item': i,
                    'Tag num.': equipment.equipment_id,
                    'Description': description_en,
                    '设备名称': equipment.name,
                    'P&ID DWG. NO.': equipment.pid_dwg_no if hasattr(equipment, 'pid_dwg_no') else '',
                    'Technical specifications': equipment.specification,
                    'QTY.': equipment.quantity if hasattr(equipment, 'quantity') else 1,
                    '单价\nPrice': equipment.unit_price if hasattr(equipment, 'unit_price') else '',
                    '总价\nTotal': equipment.total_price if hasattr(equipment, 'total_price') else '',
                    'Design tem.<br>℃': equipment.design_temperature,
                    'Design pressure<br>MPa·G': equipment.design_pressure,
                    'Operating tem.<br>℃': equipment.operating_temperature if hasattr(equipment, 'operating_temperature') else '',
                    'Operating pressure<br>MPa·G': equipment.operating_pressure if hasattr(equipment, 'operating_pressure') else '',
                    'Estimated power(kW)': equipment.estimated_power if hasattr(equipment, 'estimated_power') else '',
                    'Material': equipment.material if hasattr(equipment, 'material') else '',
                    'Insulation': equipment.insulation if hasattr(equipment, 'insulation') else '',
                    'Weight estimate(t)': equipment.weight_estimate if hasattr(equipment, 'weight_estimate') else '',
                    'Dynamic': equipment.dynamic if hasattr(equipment, 'dynamic') else '',
                    'Remark': equipment.notes
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            
            # 设置列顺序
            columns = [
                'Item', 'Tag num.', 'Description', '设备名称', 'P&ID DWG. NO.', 
                'Technical specifications', 'QTY.', '单价\nPrice', '总价\nTotal',
                'Design tem.<br>℃', 'Design pressure<br>MPa·G', 'Operating tem.<br>℃', 
                'Operating pressure<br>MPa·G', 'Estimated power(kW)', 
                'Material', 'Insulation', 'Weight estimate(t)', 
                'Dynamic', 'Remark'
            ]
            
            df = df.reindex(columns=[col for col in columns if col in df.columns])
            
            # 保存文件
            if file_path.endswith('.xlsx'):
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
            elif file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
            elif file_path.endswith('.json'):
                df.to_json(file_path, orient='records', force_ascii=False, indent=2)
            
            QMessageBox.information(self, "导出成功", f"成功导出 {len(equipment_list)} 个设备")
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出文件时发生错误:\n{str(e)}")
    
    def add_equipment(self):
        """添加设备"""
        dialog = EquipmentDialog(self)
        if dialog.exec() == QDialog.Accepted:
            equipment = dialog.get_equipment()
            if equipment and self.process_manager:
                try:
                    if self.process_manager.add_equipment(equipment):
                        self.load_equipment()
                        self.equipment_list_updated.emit()
                        self.status_bar.setText(f"设备 '{equipment.name}' 添加成功")
                    else:
                        QMessageBox.warning(self, "错误", "设备添加失败，可能设备ID已存在。")
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"添加设备时发生错误:\n{str(e)}")
    
    def edit_equipment(self):
        """编辑设备"""
        selected_row = self.equipment_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择要编辑的设备")
            return
        
        equipment_id = self.equipment_table.item(selected_row, 1).text()
        if not self.process_manager:
            return
        
        equipment = self.process_manager.get_equipment(equipment_id)
        if not equipment:
            QMessageBox.warning(self, "错误", "设备未找到")
            return
        
        dialog = EquipmentDialog(self, equipment)
        if dialog.exec() == QDialog.Accepted:
            updated_equipment = dialog.get_equipment()
            if updated_equipment and self.process_manager:
                if self.process_manager.update_equipment(updated_equipment):
                    self.load_equipment()
                    self.equipment_list_updated.emit()
                    self.status_bar.setText(f"设备 '{updated_equipment.name}' 更新成功")
                else:
                    QMessageBox.warning(self, "错误", "设备更新失败")
    
    def delete_equipment(self):
        """删除设备"""
        selected_row = self.equipment_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择要删除的设备")
            return
        
        equipment_id = self.equipment_table.item(selected_row, 1).text()
        equipment_name = self.equipment_table.item(selected_row, 2).text()
        
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除设备 '{equipment_name}' 吗？\n此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes and self.process_manager:
            if self.process_manager.delete_equipment(equipment_id):
                self.load_equipment()
                self.equipment_list_updated.emit()
                self.status_bar.setText(f"设备 '{equipment_name}' 删除成功")
            else:
                QMessageBox.warning(self, "错误", "设备删除失败")
    
    def manage_name_mapping(self):
        """管理设备名称对照表"""
        dialog = QDialog(self)
        dialog.setWindowTitle("设备名称对照表管理")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(dialog)
        
        # 添加新的对照
        add_layout = QHBoxLayout()
        add_layout.addWidget(QLabel("中文名称:"))
        chinese_input = QLineEdit()
        add_layout.addWidget(chinese_input)
        
        add_layout.addWidget(QLabel("英文名称:"))
        english_input = QLineEdit()
        add_layout.addWidget(english_input)
        
        add_btn = QPushButton("添加")
        
        def add_mapping():
            chinese = chinese_input.text().strip()
            english = english_input.text().strip()
            if chinese and english:
                self.data_manager.add_equipment_name_mapping(chinese, english)
                chinese_input.clear()
                english_input.clear()
                load_mapping_table()
        
        add_btn.clicked.connect(add_mapping)
        add_layout.addWidget(add_btn)
        
        layout.addLayout(add_layout)
        
        # 对照表表格
        mapping_table = QTableWidget()
        mapping_table.setColumnCount(2)
        mapping_table.setHorizontalHeaderLabels(["中文名称", "英文名称"])
        mapping_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(mapping_table)
        
        def load_mapping_table():
            """加载对照表到表格"""
            mapping = self.data_manager.get_equipment_name_mapping()
            mapping_table.setRowCount(len(mapping))
            
            for i, (chinese, english) in enumerate(mapping.items()):
                mapping_table.setItem(i, 0, QTableWidgetItem(chinese))
                mapping_table.setItem(i, 1, QTableWidgetItem(english))
        
        # 按钮
        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(load_mapping_table)
        btn_layout.addWidget(refresh_btn)
        
        delete_btn = QPushButton("删除选中")
        
        def delete_selected_mapping():
            """删除选中的对照表条目"""
            selected_rows = set(index.row() for index in mapping_table.selectedIndexes())
            
            for row in sorted(selected_rows, reverse=True):
                chinese_name = mapping_table.item(row, 0).text()
                self.data_manager.remove_equipment_name_mapping(chinese_name)
            
            load_mapping_table()
        
        delete_btn.clicked.connect(delete_selected_mapping)
        btn_layout.addWidget(delete_btn)
        
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        # 加载数据
        load_mapping_table()
        
        dialog.exec()


# ==================== 模板相关类 ====================

class EquipmentTemplateCreator:
    """设备清单模板创建器"""
    
    def __init__(self):
        self.templates_dir = "templates"
        os.makedirs(self.templates_dir, exist_ok=True)
        
        # 模板定义
        self.template_definitions = {
            "标准模板": {
                "description": "包含完整信息的标准设备清单模板",
                "data_start_row": 10,
                "columns": [
                    # (列名, 字段名, 宽度, 样式)
                    ("序号", "item", 6, "number"),
                    ("设备编号", "equipment_id", 12, "text"),
                    ("英文描述", "description_en", 20, "text"),
                    ("设备名称", "name", 20, "text"),
                    ("技术规格", "specification", 25, "text"),
                    ("数量", "quantity", 8, "number"),
                    ("单价", "unit_price", 12, "currency"),
                    ("总价", "total_price", 12, "currency_formula"),
                    ("设计温度℃", "design_temperature", 12, "number"),
                    ("设计压力MPa", "design_pressure", 12, "number"),
                    ("操作温度℃", "operating_temperature", 12, "number"),
                    ("操作压力MPa", "operating_pressure", 12, "number"),
                    ("估计功率kW", "estimated_power", 12, "number"),
                    ("材质", "material", 10, "text"),
                    ("保温", "insulation", 10, "text"),
                    ("重量估计t", "weight_estimate", 12, "number"),
                    ("动态", "dynamic", 10, "text"),
                    ("备注", "notes", 20, "text")
                ]
            },
            "简约模板": {
                "description": "只包含基本信息的简约模板",
                "data_start_row": 5,
                "columns": [
                    ("序号", "item", 6, "number"),
                    ("设备编号", "equipment_id", 15, "text"),
                    ("设备名称", "name", 25, "text"),
                    ("规格型号", "specification", 30, "text"),
                    ("数量", "quantity", 8, "number"),
                    ("备注", "notes", 20, "text")
                ]
            }
        }
    
    def create_template(self, template_type, file_path):
        """创建指定类型的模板"""
        if template_type not in self.template_definitions:
            template_type = "标准模板"
        
        template_def = self.template_definitions[template_type]
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "设备清单"
            
            # 创建模板
            if template_type == "标准模板":
                self._create_standard_template(ws, template_def)
            elif template_type == "简约模板":
                self._create_simple_template(ws, template_def)
            else:
                self._create_standard_template(ws, template_def)
            
            # 保存文件
            wb.save(file_path)
            print(f"模板已成功创建: {file_path}")
            return file_path
            
        except Exception as e:
            print(f"创建模板失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _create_standard_template(self, ws, template_def):
        """创建标准模板"""
        # 1. 标题
        ws.merge_cells('A1:R1')
        ws['A1'] = "工艺设备清单"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 2. 项目信息区域（使用占位符）- 简化的修复方案
        # 直接按固定位置填充，避免复杂的循环逻辑
        project_info = [
            (3, "项目名称:", "{{project_name}}", "项目编号:", "{{project_code}}"),
            (4, "项目地点:", "{{project_location}}", "编制日期:", "{{date}}"),
            (5, "设计阶段:", "{{design_phase}}", "专    业:", "{{department}}"),
            (6, "编    制:", "{{prepared_by}}", "校    对:", "{{checked_by}}"),
            (7, "审    核:", "{{approved_by}}", "版    次:", "{{revision}}")
        ]
        
        for row_num, label1, value1, label2, value2 in project_info:
            # 第一组：标签+值
            ws.cell(row=row_num, column=1, value=label1)
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value1)
            
            # 第二组：标签+值
            ws.cell(row=row_num, column=4, value=label2)
            ws.cell(row=row_num, column=4).font = Font(bold=True)
            ws.cell(row=row_num, column=5, value=value2)
        
        # 3. 表头
        header_row = 8
        columns = template_def["columns"]
        
        for i, (col_name, field, width, style) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=i, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 4. 数据区域样式
        data_start_row = template_def["data_start_row"]
        for i in range(data_start_row, data_start_row + 10):  # 10行示例数据
            for j in range(1, len(columns) + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置对齐方式
                if columns[j-1][3] == "number":
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 5. 汇总行
        summary_row = data_start_row + 12
        
        # 找到总价列的索引
        total_price_index = None
        for i, (col_name, field, width, style) in enumerate(columns, start=1):
            if field == "total_price":
                total_price_index = i
                break
        
        # 只在找到总价列时才创建汇总行
        if total_price_index and total_price_index >= 2:
            # 计算合并的列数
            merge_end_col = min(total_price_index - 2, 4)  # 最多合并到D列
            if merge_end_col >= 1:
                ws.merge_cells(f'A{summary_row}:{get_column_letter(merge_end_col)}{summary_row}')
            
            ws.cell(row=summary_row, column=1, value="合计:")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            # 总价汇总
            total_price_col = get_column_letter(total_price_index)
            ws.cell(row=summary_row, column=total_price_index-1, value="总金额:")
            ws.cell(row=summary_row, column=total_price_index, 
                value=f"=SUM({total_price_col}{data_start_row}:{total_price_col}{data_start_row+9})")
            ws.cell(row=summary_row, column=total_price_index).number_format = '#,##0.00'
        else:
            # 如果没有总价列，创建简单的汇总
            ws.merge_cells(f'A{summary_row}:D{summary_row}')
            ws.cell(row=summary_row, column=1, value="合计:")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    def _create_simple_template(self, ws, template_def):
        """创建简约模板"""
        # 1. 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = "设备清单（简约版）"
        ws['A1'].font = Font(name="微软雅黑", size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 2. 项目信息
        ws['A2'] = "项目名称: {{project_name}}"
        ws['B2'] = "项目编号: {{project_code}}"
        ws['C2'] = "日期: {{date}}"
        
        # 3. 表头
        header_row = 4
        columns = template_def["columns"]
        
        for i, (col_name, field, width, style) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=i, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 4. 数据区域样式
        data_start_row = template_def["data_start_row"]
        for i in range(data_start_row, data_start_row + 10):
            for j in range(1, len(columns) + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def create_template_config(self, template_type):
        """创建模板配置文件"""
        config = {
            "template_type": template_type,
            "created_date": datetime.now().isoformat(),
            "version": "1.0",
            "description": self.template_definitions.get(template_type, {}).get("description", ""),
            "data_start_row": self.template_definitions.get(template_type, {}).get("data_start_row", 10),
            "columns": []
        }
        
        if template_type in self.template_definitions:
            for col_name, field, width, style in self.template_definitions[template_type]["columns"]:
                config["columns"].append({
                    "display_name": col_name,
                    "field_name": field,
                    "width": width,
                    "style": style
                })
        
        config_path = os.path.join(self.templates_dir, f"{template_type}_config.json")
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        print(f"配置文件已创建: {config_path}")
        return config_path


class EquipmentTemplateFiller:
    """设备清单模板填充器"""
    
    def __init__(self):
        self.placeholders = {}
    
    def fill_template(self, template_path, output_path, equipment_list, project_info):
        """填充模板"""
        try:
            # 加载模板
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 1. 替换项目信息占位符
            self._replace_placeholders(ws, project_info)
            
            # 2. 查找数据起始行
            data_start_row = self._find_data_start_row(ws)
            
            if data_start_row is None:
                data_start_row = 10  # 默认值
            
            # 3. 解析表头，建立字段映射
            field_mapping = self._parse_header(ws, data_start_row - 1)
            
            # 4. 填充设备数据
            self._fill_equipment_data(ws, data_start_row, field_mapping, equipment_list)
            
            # 5. 更新公式
            self._update_formulas(ws, data_start_row, len(equipment_list))
            
            # 6. 保存文件
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"填充模板失败: {e}")
            return False
    
    def _replace_placeholders(self, ws, project_info):
        """替换占位符"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, value in project_info.items():
                        placeholder = "{{" + key + "}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(value))
    
    def _find_data_start_row(self, ws):
        """查找数据起始行（表头下一行）"""
        # 查找常见的表头关键词
        header_keywords = ["序号", "设备编号", "Tag num.", "设备名称", "Description"]
        
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value in header_keywords:
                    return cell.row + 1  # 数据从表头下一行开始
        
        return None
    
    def _parse_header(self, ws, header_row):
        """解析表头，建立字段映射"""
        field_mapping = {}
        
        # 列名到字段名的映射
        column_mappings = {
            "序号": "item",
            "Item": "item",
            "设备编号": "equipment_id",
            "Tag num.": "equipment_id",
            "英文描述": "description_en",
            "Description": "description_en",
            "设备名称": "name",
            "技术规格": "specification",
            "Technical specifications": "specification",
            "数量": "quantity",
            "QTY.": "quantity",
            "单价": "unit_price",
            "Price": "unit_price",
            "总价": "total_price",
            "Total": "total_price",
            "设计温度": "design_temperature",
            "Design tem.": "design_temperature",
            "设计压力": "design_pressure",
            "Design pressure": "design_pressure",
            "操作温度": "operating_temperature",
            "Operating tem.": "operating_temperature",
            "操作压力": "operating_pressure",
            "Operating pressure": "operating_pressure",
            "估计功率": "estimated_power",
            "Estimated power": "estimated_power",
            "材质": "material",
            "Material": "material",
            "保温": "insulation",
            "Insulation": "insulation",
            "重量估计": "weight_estimate",
            "Weight estimate": "weight_estimate",
            "动态": "dynamic",
            "Dynamic": "dynamic",
            "备注": "notes",
            "Remark": "notes",
        }
        
        # 遍历表头行
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            if cell.value:
                header_text = str(cell.value).strip()
                # 移除可能的换行符
                header_text = header_text.replace('\n', '').replace('<br>', '')
                
                # 查找映射
                for col_name, field_name in column_mappings.items():
                    if col_name in header_text:
                        field_mapping[field_name] = col
                        break
        
        return field_mapping
    
    def _fill_equipment_data(self, ws, start_row, field_mapping, equipment_list):
        """填充设备数据"""
        for i, equipment in enumerate(equipment_list, start=0):
            row = start_row + i
            
            # 序号
            if "item" in field_mapping:
                ws.cell(row=row, column=field_mapping["item"], value=i + 1)
            
            # 设备编号
            if "equipment_id" in field_mapping:
                ws.cell(row=row, column=field_mapping["equipment_id"], value=equipment.equipment_id)
            
            # 设备名称
            if "name" in field_mapping:
                ws.cell(row=row, column=field_mapping["name"], value=equipment.name)
            
            # 英文描述
            if "description_en" in field_mapping:
                description_en = ""
                if hasattr(equipment, 'description_en') and equipment.description_en:
                    description_en = equipment.description_en
                ws.cell(row=row, column=field_mapping["description_en"], value=description_en)
            
            # 技术规格
            if "specification" in field_mapping:
                ws.cell(row=row, column=field_mapping["specification"], value=equipment.specification)
            
            # 数量
            if "quantity" in field_mapping:
                quantity = getattr(equipment, 'quantity', 1)
                ws.cell(row=row, column=field_mapping["quantity"], value=quantity)
            
            # 单价
            if "unit_price" in field_mapping:
                unit_price = getattr(equipment, 'unit_price', 0)
                if unit_price:
                    cell = ws.cell(row=row, column=field_mapping["unit_price"], value=unit_price)
                    cell.number_format = '#,##0.00'
            
            # 总价（如果有公式就设公式，否则计算）
            if "total_price" in field_mapping and "quantity" in field_mapping and "unit_price" in field_mapping:
                qty_col = get_column_letter(field_mapping["quantity"])
                price_col = get_column_letter(field_mapping["unit_price"])
                total_col = get_column_letter(field_mapping["total_price"])
                
                formula = f"={qty_col}{row}*{price_col}{row}"
                ws.cell(row=row, column=field_mapping["total_price"], value=formula)
            
            # 设计参数
            if "design_temperature" in field_mapping:
                value = getattr(equipment, 'design_temperature', '')
                ws.cell(row=row, column=field_mapping["design_temperature"], value=value)
            
            if "design_pressure" in field_mapping:
                value = getattr(equipment, 'design_pressure', '')
                ws.cell(row=row, column=field_mapping["design_pressure"], value=value)
            
            # 备注
            if "notes" in field_mapping:
                ws.cell(row=row, column=field_mapping["notes"], value=equipment.notes)
    
    def _update_formulas(self, ws, start_row, data_count):
        """更新汇总公式"""
        if data_count == 0:
            return
        
        end_row = start_row + data_count - 1
        
        # 查找可能的汇总行（从数据结束行开始向下找10行）
        for row in range(end_row + 1, end_row + 11):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and "SUM(" in cell.value:
                    # 更新SUM公式的范围
                    old_formula = cell.value
                    # 这里需要更复杂的公式解析和更新，简化处理
                    new_formula = old_formula.replace(f"SUM(", f"SUM({start_row}:{end_row}")
                    cell.value = new_formula


# ==================== 对话框类 ====================

class ProjectInfoDialog(QDialog):
    """项目信息对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("填写项目信息")
        self.setMinimumWidth(500)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 基本信息组
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout(basic_group)
        
        self.project_name_input = QLineEdit()
        self.project_name_input.setText("某化工项目")
        
        self.project_code_input = QLineEdit()
        self.project_code_input.setText(f"PROJ-{datetime.now().strftime('%Y%m')}-001")
        
        self.project_location_input = QLineEdit()
        self.project_location_input.setText("山东")
        
        basic_layout.addRow("项目名称:", self.project_name_input)
        basic_layout.addRow("项目编号:", self.project_code_input)
        basic_layout.addRow("项目地点:", self.project_location_input)
        
        layout.addWidget(basic_group)
        
        # 设计信息组
        design_group = QGroupBox("设计信息")
        design_layout = QFormLayout(design_group)
        
        self.design_phase_combo = QComboBox()
        self.design_phase_combo.addItems(["可行性研究", "初步设计", "详细设计", "施工图设计"])
        self.design_phase_combo.setCurrentText("详细设计")
        
        self.department_combo = QComboBox()
        self.department_combo.addItems(["工艺", "管道", "设备", "电气", "仪表", "土建"])
        self.department_combo.setCurrentText("工艺")
        
        self.revision_input = QLineEdit()
        self.revision_input.setText("A")
        
        design_layout.addRow("设计阶段:", self.design_phase_combo)
        design_layout.addRow("专业:", self.department_combo)
        design_layout.addRow("版次:", self.revision_input)
        
        layout.addWidget(design_group)
        
        # 人员信息组
        person_group = QGroupBox("人员信息")
        person_layout = QFormLayout(person_group)
        
        self.prepared_by_input = QLineEdit()
        self.prepared_by_input.setPlaceholderText("编制人")
        
        self.checked_by_input = QLineEdit()
        self.checked_by_input.setPlaceholderText("校对")
        
        self.approved_by_input = QLineEdit()
        self.approved_by_input.setPlaceholderText("审核")
        
        person_layout.addRow("编制:", self.prepared_by_input)
        person_layout.addRow("校对:", self.checked_by_input)
        person_layout.addRow("审核:", self.approved_by_input)
        
        layout.addWidget(person_group)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_project_info(self):
        """获取项目信息"""
        return {
            "project_name": self.project_name_input.text().strip(),
            "project_code": self.project_code_input.text().strip(),
            "project_location": self.project_location_input.text().strip(),
            "date": datetime.now().strftime("%Y年%m月%d日"),
            "design_phase": self.design_phase_combo.currentText(),
            "department": self.department_combo.currentText(),
            "revision": self.revision_input.text().strip(),
            "prepared_by": self.prepared_by_input.text().strip() or "系统生成",
            "checked_by": self.checked_by_input.text().strip(),
            "approved_by": self.approved_by_input.text().strip()
        }


class TemplateTypeDialog(QDialog):
    """模板类型选择对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择模板类型")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("请选择要创建的模板类型:"))
        
        self.template_type_combo = QComboBox()
        self.template_type_combo.addItems([
            "标准模板",
            "简约模板"
        ])
        
        layout.addWidget(self.template_type_combo)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_template_type(self):
        return self.template_type_combo.currentText()


class AdvancedSearchDialog(QDialog):
    """高级搜索对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("高级搜索设备")
        self.setMinimumWidth(400)
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        
        # 制造商搜索
        self.manufacturer_input = QLineEdit()
        self.manufacturer_input.setPlaceholderText("制造商名称")
        form_layout.addRow("制造商:", self.manufacturer_input)
        
        # 安装位置搜索
        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("安装位置")
        form_layout.addRow("安装位置:", self.location_input)
        
        # 投用日期范围
        date_layout = QHBoxLayout()
        self.start_date = QLineEdit()
        self.start_date.setPlaceholderText("YYYY-MM-DD")
        self.end_date = QLineEdit()
        self.end_date.setPlaceholderText("YYYY-MM-DD")
        date_layout.addWidget(self.start_date)
        date_layout.addWidget(QLabel("至"))
        date_layout.addWidget(self.end_date)
        form_layout.addRow("投用日期范围:", date_layout)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_search_criteria(self):
        """获取搜索条件"""
        criteria = {}
        
        manufacturer = self.manufacturer_input.text().strip()
        if manufacturer:
            criteria['manufacturer'] = manufacturer
        
        location = self.location_input.text().strip()
        if location:
            criteria['location'] = location
        
        start_date = self.start_date.text().strip()
        if start_date:
            criteria['start_date'] = start_date
        
        end_date = self.end_date.text().strip()
        if end_date:
            criteria['end_date'] = end_date
        
        return criteria


class BatchEditDialog(QDialog):
    """批量编辑对话框"""
    
    def __init__(self, equipment_ids, process_manager, parent=None):
        super().__init__(parent)
        self.equipment_ids = equipment_ids
        self.process_manager = process_manager
        self.setWindowTitle("批量编辑设备")
        self.setMinimumWidth(400)
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel(f"批量编辑 {len(self.equipment_ids)} 个设备"))
        
        form_layout = QFormLayout()
        
        # 可批量编辑的字段
        self.status_combo = QComboBox()
        self.status_combo.addItems(["", "运行中", "停机", "维修中", "备用"])
        self.status_combo.setCurrentIndex(0)
        form_layout.addRow("状态:", self.status_combo)
        
        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("留空则不修改")
        form_layout.addRow("安装位置:", self.location_input)
        
        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(100)
        self.notes_input.setPlaceholderText("留空则不修改")
        form_layout.addRow("备注:", self.notes_input)
        
        layout.addLayout(form_layout)
        
        layout.addWidget(QLabel("注意：空字段不会修改原有值"))
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.apply_changes)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def apply_changes(self):
        """应用批量修改"""
        changes = {}
        
        status = self.status_combo.currentText()
        if status:
            changes['status'] = status
        
        location = self.location_input.text().strip()
        if location:
            changes['location'] = location
        
        notes = self.notes_input.toPlainText().strip()
        if notes:
            changes['notes'] = notes
        
        if not changes:
            QMessageBox.warning(self, "警告", "没有修改任何字段")
            return
        
        success_count = 0
        for equipment_id in self.equipment_ids:
            equipment = self.process_manager.get_equipment(equipment_id)
            if equipment:
                # 应用修改
                if 'status' in changes:
                    equipment.status = changes['status']
                if 'location' in changes:
                    equipment.location = changes['location']
                if 'notes' in changes:
                    equipment.notes = changes['notes']
                
                if self.process_manager.update_equipment(equipment):
                    success_count += 1
        
        QMessageBox.information(self, "批量编辑完成", f"成功更新 {success_count} 个设备")
        self.accept()


class EquipmentDialog(QDialog):
    """设备对话框 - 改进版，根据设备类型动态显示规格字段"""
    
    def __init__(self, parent=None, equipment=None):
        super().__init__(parent)
        self.parent_widget = parent
        self.equipment = equipment
        self.setWindowTitle("添加设备" if not equipment else "编辑设备")
        self.setMinimumWidth(700)
        
        # 存储规格参数的字典
        self.spec_params = {}
        # 存储体积相关参数
        self.volume_widgets = {}
        
        self.setup_ui()
        
        # 如果是编辑模式，先加载数据再设置类型
        if equipment:
            # 保存设备类型
            equipment_type = equipment.equipment_type if hasattr(equipment, 'equipment_type') else equipment.type
            # 设置设备类型
            index = self.type_combo.findText(equipment_type)
            if index >= 0:
                self.type_combo.setCurrentIndex(index)
            
            # 设置设备类型相关的字段
            self.on_type_changed(equipment_type)
            
            # 然后加载其他数据
            self.load_equipment_data()
        else:
            # 添加模式：初始设置设备类型相关的字段
            self.on_type_changed(self.type_combo.currentText())
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        
        # 使用标签页组织内容
        tab_widget = QTabWidget()
        
        # ==================== 基本信息标签页 ====================
        basic_tab = QWidget()
        basic_layout = QFormLayout(basic_tab)
        
        self.equipment_id_input = QLineEdit()
        self.name_input = QLineEdit()
        
        # 添加英文描述输入框
        self.description_en_input = QLineEdit()
        self.description_en_input.setPlaceholderText("自动从对照表获取，可修改")
        
        self.type_combo = QComboBox()
        self.type_combo.addItems(["反应器", "分离器", "换热器", "泵", "压缩机", "储罐", "阀门", "管道", "其他"])
        self.type_combo.currentTextChanged.connect(self.on_type_changed)
        
        self.model_input = QLineEdit()
        self.manufacturer_input = QLineEdit()
        self.location_input = QLineEdit()
        
        # 状态信息合并到基本信息
        self.status_combo = QComboBox()
        self.status_combo.addItems(["运行中", "停机", "维修中", "备用"])
        
        self.commission_date = QLineEdit()
        self.commission_date.setPlaceholderText("YYYY-MM-DD")
        
        basic_layout.addRow("设备ID*:", self.equipment_id_input)
        basic_layout.addRow("设备名称(中文)*:", self.name_input)
        basic_layout.addRow("设备名称(英文):", self.description_en_input)
        basic_layout.addRow("设备类型:", self.type_combo)
        basic_layout.addRow("型号:", self.model_input)
        basic_layout.addRow("制造商:", self.manufacturer_input)
        basic_layout.addRow("安装位置:", self.location_input)
        basic_layout.addRow("状态:", self.status_combo)
        basic_layout.addRow("投用日期:", self.commission_date)
        
        # 连接信号：当中文名称变化时，自动获取英文名称
        self.name_input.textChanged.connect(self.update_english_name)
        
        tab_widget.addTab(basic_tab, "基本信息")
        
        # ==================== 技术规格标签页 ====================
        self.spec_tab = QWidget()
        self.spec_layout = QVBoxLayout(self.spec_tab)
        
        # 滚动区域用于容纳大量规格字段
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        self.spec_form_layout = QFormLayout(scroll_content)
        scroll_area.setWidget(scroll_content)
        self.spec_layout.addWidget(scroll_area)
        
        # 添加规格参数输入框
        self.setup_specification_fields()
        
        tab_widget.addTab(self.spec_tab, "技术规格")
        
        # ==================== 设计操作参数标签页 ====================
        design_operation_tab = QWidget()
        design_operation_layout = QVBoxLayout(design_operation_tab)
        
        # 设计参数组
        design_group = QGroupBox("设计参数")
        design_layout = QFormLayout(design_group)
        
        # 设计压力 - 增加常温常压选项
        design_pressure_layout = QHBoxLayout()
        self.design_pressure_input = QDoubleSpinBox()
        self.design_pressure_input.setRange(0, 1000)
        self.design_pressure_input.setDecimals(3)
        self.design_pressure_input.setSuffix(" MPa")
        self.design_pressure_input.setMinimumWidth(100)
        
        self.normal_pressure_check = QCheckBox("常压 (0.1013 MPa)")
        self.normal_pressure_check.toggled.connect(self.on_normal_pressure_toggled)
        
        design_pressure_layout.addWidget(self.design_pressure_input)
        design_pressure_layout.addWidget(self.normal_pressure_check)
        design_pressure_layout.addStretch()
        
        design_layout.addRow("设计压力:", design_pressure_layout)
        
        # 设计温度 - 增加常温选项
        design_temperature_layout = QHBoxLayout()
        self.design_temperature_input = QDoubleSpinBox()
        self.design_temperature_input.setRange(-273, 1000)
        self.design_temperature_input.setDecimals(1)
        self.design_temperature_input.setSuffix(" °C")
        self.design_temperature_input.setMinimumWidth(100)
        
        self.normal_temperature_check = QCheckBox("常温 (25 °C)")
        self.normal_temperature_check.toggled.connect(self.on_normal_temperature_toggled)
        
        design_temperature_layout.addWidget(self.design_temperature_input)
        design_temperature_layout.addWidget(self.normal_temperature_check)
        design_temperature_layout.addStretch()
        
        design_layout.addRow("设计温度:", design_temperature_layout)
        
        design_operation_layout.addWidget(design_group)
        
        # 操作参数组
        operation_group = QGroupBox("操作参数")
        operation_layout = QFormLayout(operation_group)
        
        # 操作压力
        operating_pressure_layout = QHBoxLayout()
        self.operating_pressure_input = QDoubleSpinBox()
        self.operating_pressure_input.setRange(0, 1000)
        self.operating_pressure_input.setDecimals(3)
        self.operating_pressure_input.setSuffix(" MPa")
        self.operating_pressure_input.setMinimumWidth(100)
        
        self.operating_normal_pressure_check = QCheckBox("常压 (0.1013 MPa)")
        self.operating_normal_pressure_check.toggled.connect(self.on_operating_normal_pressure_toggled)
        
        operating_pressure_layout.addWidget(self.operating_pressure_input)
        operating_pressure_layout.addWidget(self.operating_normal_pressure_check)
        operating_pressure_layout.addStretch()
        
        operation_layout.addRow("操作压力:", operating_pressure_layout)
        
        # 操作温度
        operating_temperature_layout = QHBoxLayout()
        self.operating_temperature_input = QDoubleSpinBox()
        self.operating_temperature_input.setRange(-273, 1000)
        self.operating_temperature_input.setDecimals(1)
        self.operating_temperature_input.setSuffix(" °C")
        self.operating_temperature_input.setMinimumWidth(100)
        
        self.operating_normal_temperature_check = QCheckBox("常温 (25 °C)")
        self.operating_normal_temperature_check.toggled.connect(self.on_operating_normal_temperature_toggled)
        
        operating_temperature_layout.addWidget(self.operating_temperature_input)
        operating_temperature_layout.addWidget(self.operating_normal_temperature_check)
        operating_temperature_layout.addStretch()
        
        operation_layout.addRow("操作温度:", operating_temperature_layout)
        
        design_operation_layout.addWidget(operation_group)
        
        # 电机功率组
        power_group = QGroupBox("电机功率")
        power_layout = QFormLayout(power_group)
        
        # 设备数量
        self.quantity_input = QSpinBox()
        self.quantity_input.setRange(1, 1000)
        self.quantity_input.setValue(1)
        self.quantity_input.valueChanged.connect(self.on_quantity_changed)
        power_layout.addRow("设备数量:", self.quantity_input)
        
        # 运行设备数量
        self.running_quantity_input = QSpinBox()
        self.running_quantity_input.setRange(1, 1000)
        self.running_quantity_input.setValue(1)
        self.running_quantity_input.valueChanged.connect(self.on_running_quantity_changed)
        power_layout.addRow("运行设备数量:", self.running_quantity_input)
        
        # 单机功率
        self.single_power_input = QDoubleSpinBox()
        self.single_power_input.setRange(0, 10000)
        self.single_power_input.setDecimals(2)
        self.single_power_input.setSuffix(" kW")
        self.single_power_input.valueChanged.connect(self.on_single_power_changed)
        power_layout.addRow("单机功率:", self.single_power_input)
        
        # 运行功率
        self.operating_power_input = QDoubleSpinBox()
        self.operating_power_input.setRange(0, 10000)
        self.operating_power_input.setDecimals(2)
        self.operating_power_input.setSuffix(" kW")
        power_layout.addRow("运行功率:", self.operating_power_input)
        
        # 总功率
        self.total_power_input = QDoubleSpinBox()
        self.total_power_input.setRange(0, 100000)
        self.total_power_input.setDecimals(2)
        self.total_power_input.setSuffix(" kW")
        power_layout.addRow("总功率:", self.total_power_input)
        
        # 电机参数
        power_layout.addRow(QLabel("电机参数:"))
        
        # 是否变频
        self.frequency_conversion_check = QCheckBox("变频")
        power_layout.addRow("", self.frequency_conversion_check)
        
        design_operation_layout.addWidget(power_group)
        design_operation_layout.addStretch()
        
        tab_widget.addTab(design_operation_tab, "设计操作参数")
        
        # ==================== 导出与其他标签页 ====================
        export_other_tab = QWidget()
        export_other_layout = QVBoxLayout(export_other_tab)
        
        # 导出参数组
        export_group = QGroupBox("导出参数")
        export_layout = QFormLayout(export_group)
        
        self.pid_dwg_no_input = QLineEdit()
        self.pid_dwg_no_input.setPlaceholderText("P&ID图号")
        
        self.unit_price_input = QDoubleSpinBox()
        self.unit_price_input.setRange(0, 10000000)
        self.unit_price_input.setDecimals(2)
        self.unit_price_input.setPrefix("¥ ")
        
        self.total_price_input = QDoubleSpinBox()
        self.total_price_input.setRange(0, 10000000)
        self.total_price_input.setDecimals(2)
        self.total_price_input.setPrefix("¥ ")
        
        self.material_input = QLineEdit()
        self.material_input.setPlaceholderText("材质")
        
        self.insulation_input = QLineEdit()
        self.insulation_input.setPlaceholderText("保温")
        
        self.weight_estimate_input = QDoubleSpinBox()
        self.weight_estimate_input.setRange(0, 1000)
        self.weight_estimate_input.setDecimals(2)
        self.weight_estimate_input.setSuffix(" t")
        
        self.dynamic_input = QLineEdit()
        self.dynamic_input.setPlaceholderText("动态")
        
        export_layout.addRow("P&ID图号:", self.pid_dwg_no_input)
        export_layout.addRow("单价:", self.unit_price_input)
        export_layout.addRow("总价:", self.total_price_input)
        export_layout.addRow("材质:", self.material_input)
        export_layout.addRow("保温:", self.insulation_input)
        export_layout.addRow("重量估计:", self.weight_estimate_input)
        export_layout.addRow("动态:", self.dynamic_input)
        
        export_other_layout.addWidget(export_group)
        
        # 其他信息组
        other_group = QGroupBox("其他信息")
        other_layout = QVBoxLayout(other_group)
        
        self.notes_input = QTextEdit()
        self.notes_input.setPlaceholderText("备注信息...")
        other_layout.addWidget(self.notes_input)
        
        export_other_layout.addWidget(other_group)
        export_other_layout.addStretch()
        
        tab_widget.addTab(export_other_tab, "导出与其他")
        
        layout.addWidget(tab_widget)
        
        # 验证状态
        self.validation_label = QLabel()
        self.validation_label.setStyleSheet("color: red;")
        layout.addWidget(self.validation_label)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def setup_specification_fields(self):
        """设置规格参数输入框（根据设备类型动态显示）"""
        # 清空现有的规格字段
        self.clear_layout(self.spec_form_layout)
        self.spec_params = {}
        self.volume_widgets = {}
        
        # 罐类规格参数（反应器、储罐、分离器等）
        if self.type_combo.currentText() in ["储罐", "反应器", "分离器"]:
            # 直径
            self.add_spec_field("直径", "diameter", "mm")
            
            # 高度
            self.add_spec_field("高度", "height", "mm")
            
            # 体积 - 添加计算按钮
            volume_layout = QHBoxLayout()
            self.volume_input = QDoubleSpinBox()
            self.volume_input.setRange(0, 1000000)
            self.volume_input.setDecimals(2)
            self.volume_input.setSuffix(" m³")
            self.volume_input.setMinimumWidth(100)
            
            # 保存到字典
            self.spec_params['volume'] = {
                'label': '体积',
                'widget': self.volume_input,
                'unit': 'm³',
                'type': 'spinbox'
            }
            
            # 计算按钮
            calculate_btn = QPushButton("计算")
            calculate_btn.setMaximumWidth(60)
            calculate_btn.clicked.connect(self.calculate_volume)
            
            volume_layout.addWidget(self.volume_input)
            volume_layout.addWidget(calculate_btn)
            volume_layout.addStretch()
            
            self.spec_form_layout.addRow("体积:", volume_layout)
            
            # 保存直径和高度控件用于计算体积
            self.volume_widgets['diameter'] = self.spec_params.get('diameter', {}).get('widget')
            self.volume_widgets['height'] = self.spec_params.get('height', {}).get('widget')
            
            # 其他参数
            self.add_spec_field("上封头类型", "top_head_type", "")
            self.add_spec_field("下封头类型", "bottom_head_type", "")
            self.add_spec_field("壁厚", "wall_thickness", "mm")
            self.add_spec_field("材质", "material_spec", "")
            
            # 如果有搅拌功能
            if self.type_combo.currentText() == "反应器":
                self.add_spec_field("搅拌类型", "agitation_type", "")
                self.add_spec_field("搅拌转速", "agitation_speed", "rpm")
                self.add_spec_field("搅拌功率", "agitation_power", "kW")
        
        # 泵类规格参数
        elif self.type_combo.currentText() == "泵":
            self.add_spec_field("流量", "flow_rate", "m³/h")
            self.add_spec_field("扬程", "head", "m")
            self.add_spec_field("必需汽蚀余量", "npsh", "m")
            self.add_spec_field("效率", "efficiency", "%")
            self.add_spec_field("进出口直径", "port_diameter", "mm")
            self.add_spec_field("密封形式", "seal_type", "")
            self.add_spec_field("物料密度", "material_density", "kg/m³")
            self.add_spec_field("物料粘度", "material_viscosity", "cP")
            self.add_checkbox_field("变频", "frequency_conversion")
        
        # 压缩机规格参数
        elif self.type_combo.currentText() == "压缩机":
            self.add_spec_field("排气量", "displacement", "m³/min")
            self.add_spec_field("排气压力", "discharge_pressure", "MPa")
            self.add_spec_field("吸气压力", "suction_pressure", "MPa")
            self.add_spec_field("冷却方式", "cooling_method", "")
            self.add_spec_field("润滑方式", "lubrication_method", "")
        
        # 换热器规格参数
        elif self.type_combo.currentText() == "换热器":
            self.add_spec_field("换热面积", "heat_exchange_area", "m²")
            self.add_spec_field("管程数", "tube_passes", "")
            self.add_spec_field("壳程数", "shell_passes", "")
            self.add_spec_field("管径", "tube_diameter", "mm")
            self.add_spec_field("管长", "tube_length", "mm")
            self.add_spec_field("管材", "tube_material", "")
            self.add_spec_field("壳材", "shell_material", "")
        
        # 阀门规格参数
        elif self.type_combo.currentText() == "阀门":
            self.add_spec_field("公称直径", "nominal_diameter", "mm")
            self.add_spec_field("公称压力", "nominal_pressure", "MPa")
            self.add_spec_field("连接形式", "connection_type", "")
            self.add_spec_field("阀体材质", "valve_body_material", "")
            self.add_spec_field("阀座材质", "valve_seat_material", "")
            self.add_spec_field("驱动方式", "actuation_type", "")
        
        # 管道规格参数
        elif self.type_combo.currentText() == "管道":
            self.add_spec_field("公称直径", "nominal_diameter", "mm")
            self.add_spec_field("外径", "outer_diameter", "mm")
            self.add_spec_field("壁厚", "wall_thickness", "mm")
            self.add_spec_field("长度", "length", "m")
            self.add_spec_field("管道等级", "pipe_class", "")
            self.add_spec_field("连接方式", "connection_method", "")
        
        # 其他设备 - 通用规格参数
        else:
            self.add_spec_field("主要规格", "main_spec", "")
            self.add_spec_field("处理能力", "capacity", "")
            self.add_spec_field("工作介质", "working_medium", "")
        
        # 其他规格输入框（所有设备类型都有）
        self.add_textedit_field("其他规格", "other_specifications")
    
    def calculate_volume(self):
        """根据直径和高度计算体积"""
        diameter_widget = self.volume_widgets.get('diameter')
        height_widget = self.volume_widgets.get('height')
        
        if diameter_widget and height_widget:
            diameter = diameter_widget.value()  # mm
            height = height_widget.value()  # mm
            
            if diameter > 0 and height > 0:
                # 将直径和高度转换为米
                diameter_m = diameter / 1000.0
                height_m = height / 1000.0
                
                # 计算圆柱体体积: V = π * (d/2)² * h
                radius = diameter_m / 2.0
                volume = 3.141592653589793 * radius * radius * height_m
                
                # 设置体积值
                self.volume_input.setValue(round(volume, 2))
                
                # 显示提示
                self.show_tooltip(f"根据直径 {diameter} mm 和高度 {height} mm 计算得出")
    
    def show_tooltip(self, message):
        """显示工具提示"""
        QToolTip.showText(self.mapToGlobal(QPoint(0, 0)), message, self, QRect(), 2000)
    
    def add_spec_field(self, label_text, field_name, unit):
        """添加规格字段"""
        layout = QHBoxLayout()
        
        # 输入框
        if unit in ["m³", "kW", "MPa", "m", "m³/h", "m²"]:  # 需要小数精度的单位
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 1000000)
            input_widget.setDecimals(2)
        elif unit in ["mm", "kg/m³", "cP", "m³/min"]:  # 需要整数或一位小数的单位
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 1000000)
            input_widget.setDecimals(1)
        elif unit in ["%", "rpm"]:  # 需要整数的单位
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 1000000)
            input_widget.setDecimals(0)
        else:
            input_widget = QLineEdit()
        
        # 设置单位
        if unit:
            if isinstance(input_widget, QDoubleSpinBox):
                input_widget.setSuffix(f" {unit}")
                input_widget.setMinimumWidth(100)
            else:
                input_widget.setPlaceholderText(unit)
        
        layout.addWidget(input_widget)
        layout.addStretch()
        
        # 添加到布局
        self.spec_form_layout.addRow(f"{label_text}:", layout)
        
        # 保存到字典
        self.spec_params[field_name] = {
            'label': label_text,
            'widget': input_widget,
            'unit': unit,
            'type': 'spinbox' if isinstance(input_widget, QDoubleSpinBox) else 'lineedit'
        }
    
    def add_checkbox_field(self, label_text, field_name):
        """添加复选框字段"""
        checkbox = QCheckBox()
        self.spec_form_layout.addRow(f"{label_text}:", checkbox)
        
        # 保存到字典
        self.spec_params[field_name] = {
            'label': label_text,
            'widget': checkbox,
            'type': 'checkbox'
        }
    
    def add_textedit_field(self, label_text, field_name):
        """添加多行文本字段"""
        textedit = QTextEdit()
        textedit.setMaximumHeight(100)
        self.spec_form_layout.addRow(f"{label_text}:", textedit)
        
        # 保存到字典
        self.spec_params[field_name] = {
            'label': label_text,
            'widget': textedit,
            'type': 'textedit'
        }
    
    def clear_layout(self, layout):
        """清空布局中的所有控件"""
        while layout.rowCount() > 0:
            layout.removeRow(0)
    
    def on_type_changed(self, equipment_type):
        """设备类型变化时，重新设置规格字段"""
        self.setup_specification_fields()
    
    def on_normal_pressure_toggled(self, checked):
        """设计压力常压选项切换"""
        if checked:
            self.design_pressure_input.setValue(0.1013)
            self.design_pressure_input.setEnabled(False)
        else:
            self.design_pressure_input.setEnabled(True)
    
    def on_normal_temperature_toggled(self, checked):
        """设计温度常温选项切换"""
        if checked:
            self.design_temperature_input.setValue(25.0)
            self.design_temperature_input.setEnabled(False)
        else:
            self.design_temperature_input.setEnabled(True)
    
    def on_operating_normal_pressure_toggled(self, checked):
        """操作压力常压选项切换"""
        if checked:
            self.operating_pressure_input.setValue(0.1013)
            self.operating_pressure_input.setEnabled(False)
        else:
            self.operating_pressure_input.setEnabled(True)
    
    def on_operating_normal_temperature_toggled(self, checked):
        """操作温度常温选项切换"""
        if checked:
            self.operating_temperature_input.setValue(25.0)
            self.operating_temperature_input.setEnabled(False)
        else:
            self.operating_temperature_input.setEnabled(True)
    
    def on_quantity_changed(self, value):
        """设备数量变化时，重新计算总功率"""
        self.calculate_total_power()
    
    def on_running_quantity_changed(self, value):
        """运行设备数量变化时，重新计算运行功率"""
        self.calculate_operating_power()
    
    def on_single_power_changed(self, value):
        """单机功率变化时，重新计算运行功率和总功率"""
        self.calculate_operating_power()
        self.calculate_total_power()
    
    def calculate_operating_power(self):
        """计算运行功率"""
        single_power = self.single_power_input.value()
        running_quantity = self.running_quantity_input.value()
        
        if single_power > 0 and running_quantity > 0:
            operating_power = single_power * running_quantity
            self.operating_power_input.setValue(operating_power)
    
    def calculate_total_power(self):
        """计算总功率"""
        single_power = self.single_power_input.value()
        quantity = self.quantity_input.value()
        
        if single_power > 0 and quantity > 0:
            total_power = single_power * quantity
            self.total_power_input.setValue(total_power)
    
    def update_english_name(self):
        """根据中文名称自动获取英文名称"""
        chinese_name = self.name_input.text().strip()
        if chinese_name and self.parent_widget and hasattr(self.parent_widget, 'data_manager'):
            # 尝试从对照表获取英文名称
            english_name = self.parent_widget.data_manager.get_english_name(chinese_name)
            if english_name:
                self.description_en_input.setText(english_name)
    
    def validate_and_accept(self):
        """验证并接受"""
        equipment_id = self.equipment_id_input.text().strip()
        name = self.name_input.text().strip()
        
        errors = []
        
        if not equipment_id:
            errors.append("设备ID不能为空")
        if not name:
            errors.append("设备名称不能为空")
        
        # 验证日期格式
        if self.commission_date.text().strip():
            if not re.match(r'^\d{4}-\d{2}-\d{2}$', self.commission_date.text().strip()):
                errors.append("投用日期格式应为: YYYY-MM-DD")
        
        if errors:
            self.validation_label.setText("错误: " + "; ".join(errors))
            return
        
        self.validation_label.setText("")
        self.accept()
    
    def load_equipment_data(self):
        """加载设备数据"""
        if not self.equipment:
            return
        
        # 基本信息
        self.equipment_id_input.setText(self.equipment.equipment_id)
        self.name_input.setText(self.equipment.name)
        
        index = self.type_combo.findText(self.equipment.equipment_type if hasattr(self.equipment, 'equipment_type') else self.equipment.type)
        if index >= 0:
            self.type_combo.setCurrentIndex(index)
        
        self.model_input.setText(self.equipment.model or "")
        self.manufacturer_input.setText(self.equipment.manufacturer or "")
        self.location_input.setText(self.equipment.location or "")
        
        index = self.status_combo.findText(self.equipment.status)
        if index >= 0:
            self.status_combo.setCurrentIndex(index)
        
        if hasattr(self.equipment, 'commission_date') and self.equipment.commission_date:
            if hasattr(self.equipment.commission_date, 'strftime'):
                self.commission_date.setText(self.equipment.commission_date.strftime('%Y-%m-%d'))
            else:
                self.commission_date.setText(str(self.equipment.commission_date))
        
        # 加载设计参数 - 修复：添加空值检查
        if hasattr(self.equipment, 'design_pressure') and self.equipment.design_pressure:
            if isinstance(self.equipment.design_pressure, str) and self.equipment.design_pressure == "NP":
                self.normal_pressure_check.setChecked(True)
            else:
                try:
                    pressure_val = float(self.equipment.design_pressure)
                    if abs(pressure_val - 0.1013) < 0.0001:
                        self.normal_pressure_check.setChecked(True)
                    else:
                        self.design_pressure_input.setValue(pressure_val)
                except (ValueError, TypeError):
                    self.design_pressure_input.setValue(0.0)

        if hasattr(self.equipment, 'design_temperature') and self.equipment.design_temperature:
            if isinstance(self.equipment.design_temperature, str) and self.equipment.design_temperature == "NT":
                self.normal_temperature_check.setChecked(True)
            else:
                try:
                    temp_val = float(self.equipment.design_temperature)
                    if abs(temp_val - 25.0) < 0.001:
                        self.normal_temperature_check.setChecked(True)
                    else:
                        self.design_temperature_input.setValue(temp_val)
                except (ValueError, TypeError):
                    self.design_temperature_input.setValue(0.0)
        
        # 加载操作参数 - 修复：添加空值检查
        if hasattr(self.equipment, 'operating_pressure') and self.equipment.operating_pressure:
            if isinstance(self.equipment.operating_pressure, str) and self.equipment.operating_pressure == "NP":
                self.operating_normal_pressure_check.setChecked(True)
            else:
                try:
                    pressure_val = float(self.equipment.operating_pressure)
                    if abs(pressure_val - 0.1013) < 0.0001:
                        self.operating_normal_pressure_check.setChecked(True)
                    else:
                        self.operating_pressure_input.setValue(pressure_val)
                except (ValueError, TypeError):
                    self.operating_pressure_input.setValue(0.0)
        
        if hasattr(self.equipment, 'operating_temperature') and self.equipment.operating_temperature:
            if isinstance(self.equipment.operating_temperature, str) and self.equipment.operating_temperature == "NT":
                self.operating_normal_temperature_check.setChecked(True)
            else:
                try:
                    temp_val = float(self.equipment.operating_temperature)
                    if abs(temp_val - 25.0) < 0.001:
                        self.operating_normal_temperature_check.setChecked(True)
                    else:
                        self.operating_temperature_input.setValue(temp_val)
                except (ValueError, TypeError):
                    self.operating_temperature_input.setValue(0.0)
        
        # 加载英文描述
        if hasattr(self.equipment, 'description_en') and self.equipment.description_en:
            self.description_en_input.setText(self.equipment.description_en)
        else:
            # 尝试从对照表获取
            if self.parent_widget and hasattr(self.parent_widget, 'data_manager'):
                english_name = self.parent_widget.data_manager.get_english_name(self.equipment.name)
                if english_name:
                    self.description_en_input.setText(english_name)
        
        # 加载电机功率参数 - 修复：添加空值检查
        if hasattr(self.equipment, 'quantity'):
            self.quantity_input.setValue(self.equipment.quantity or 1)
        
        if hasattr(self.equipment, 'running_quantity'):
            running_quantity = getattr(self.equipment, 'running_quantity')
            if running_quantity is not None:
                self.running_quantity_input.setValue(running_quantity)
            else:
                self.running_quantity_input.setValue(1)
        
        if hasattr(self.equipment, 'single_power'):
            single_power = getattr(self.equipment, 'single_power')
            if single_power is not None:
                self.single_power_input.setValue(float(single_power))
            else:
                self.single_power_input.setValue(0.0)
        
        if hasattr(self.equipment, 'operating_power'):
            operating_power = getattr(self.equipment, 'operating_power')
            if operating_power is not None:
                self.operating_power_input.setValue(float(operating_power))
            else:
                self.operating_power_input.setValue(0.0)
        
        if hasattr(self.equipment, 'total_power'):
            total_power = getattr(self.equipment, 'total_power')
            if total_power is not None:
                self.total_power_input.setValue(float(total_power))
            else:
                self.total_power_input.setValue(0.0)
        
        if hasattr(self.equipment, 'frequency_conversion'):
            self.frequency_conversion_check.setChecked(bool(getattr(self.equipment, 'frequency_conversion', False)))
        
        # 加载导出相关参数 - 修复：添加空值检查
        if hasattr(self.equipment, 'pid_dwg_no'):
            self.pid_dwg_no_input.setText(self.equipment.pid_dwg_no or "")
        
        if hasattr(self.equipment, 'unit_price') and self.equipment.unit_price:
            self.unit_price_input.setValue(float(self.equipment.unit_price))
        else:
            self.unit_price_input.setValue(0.0)
        
        if hasattr(self.equipment, 'total_price') and self.equipment.total_price:
            self.total_price_input.setValue(float(self.equipment.total_price))
        else:
            self.total_price_input.setValue(0.0)
        
        if hasattr(self.equipment, 'material'):
            self.material_input.setText(self.equipment.material or "")
        
        if hasattr(self.equipment, 'insulation'):
            self.insulation_input.setText(self.equipment.insulation or "")
        
        if hasattr(self.equipment, 'weight_estimate') and self.equipment.weight_estimate:
            self.weight_estimate_input.setValue(float(self.equipment.weight_estimate))
        else:
            self.weight_estimate_input.setValue(0.0)
        
        if hasattr(self.equipment, 'dynamic'):
            self.dynamic_input.setText(self.equipment.dynamic or "")
        
        self.notes_input.setText(self.equipment.notes or "")
        
        # ==================== 新增：加载技术规格数据 ====================
        # 等待界面完全加载后再解析技术规格
        QTimer.singleShot(100, self.parse_specification_data)
    
    def parse_specification_data(self):
        """解析并加载技术规格数据"""
        if not self.equipment or not self.equipment.specification:
            return
        
        # 从规格字符串中解析数据
        spec_str = self.equipment.specification
        if not spec_str:
            return
        
        # 按分号分割规格项
        spec_items = [item.strip() for item in spec_str.split(';') if item.strip()]
        
        # 解析每个规格项
        for item in spec_items:
            # 查找冒号分隔标签和值
            if ':' in item:
                parts = item.split(':', 1)
                if len(parts) == 2:
                    label = parts[0].strip()
                    value_part = parts[1].strip()
                    
                    # 查找对应的控件并设置值
                    self.set_specification_value(label, value_part)
        
        # 特殊处理：尝试从规格中提取体积相关参数
        self.extract_volume_from_spec(spec_str)
        
    def set_specification_value(self, label, value_part):
        """根据标签设置规格控件的值"""
        # 遍历所有规格参数
        for field_name, field_info in self.spec_params.items():
            field_label = field_info.get('label', '')
            
            # 检查标签是否匹配
            if field_label and field_label in label:
                widget = field_info.get('widget')
                field_type = field_info.get('type', '')
                
                if not widget:
                    continue
                
                try:
                    if field_type == 'spinbox':
                        # 提取数值部分
                        import re
                        # 查找数字（包括小数）
                        numbers = re.findall(r'[-+]?\d*\.\d+|\d+', value_part)
                        if numbers:
                            num_value = float(numbers[0])
                            widget.setValue(num_value)
                    
                    elif field_type == 'lineedit':
                        # 设置文本值
                        widget.setText(value_part)
                    
                    elif field_type == 'checkbox':
                        # 检查是否是/否
                        if '是' in value_part:
                            widget.setChecked(True)
                        elif '否' in value_part:
                            widget.setChecked(False)
                    
                    elif field_type == 'textedit':
                        # 对于文本编辑框，需要特殊处理
                        if field_name == 'other_specifications':
                            # 获取当前文本
                            current_text = widget.toPlainText()
                            if current_text:
                                widget.setPlainText(current_text + '\n' + label + ': ' + value_part)
                            else:
                                widget.setPlainText(label + ': ' + value_part)
                
                except Exception as e:
                    print(f"设置规格字段 {label} 失败: {e}")
                    
    def extract_volume_from_spec(self, spec_str):
        """从规格字符串中提取体积相关参数"""
        if not spec_str:
            return
        
        # 查找直径
        diameter_match = re.search(r'直径[:：]\s*([\d\.]+)\s*mm', spec_str)
        if diameter_match and 'diameter' in self.spec_params:
            try:
                diameter = float(diameter_match.group(1))
                diameter_widget = self.spec_params['diameter'].get('widget')
                if diameter_widget:
                    diameter_widget.setValue(diameter)
            except:
                pass
        
        # 查找高度
        height_match = re.search(r'高度[:：]\s*([\d\.]+)\s*mm', spec_str)
        if height_match and 'height' in self.spec_params:
            try:
                height = float(height_match.group(1))
                height_widget = self.spec_params['height'].get('widget')
                if height_widget:
                    height_widget.setValue(height)
            except:
                pass
        
        # 查找体积
        volume_match = re.search(r'体积[:：]\s*([\d\.]+)\s*m³', spec_str)
        if volume_match and 'volume' in self.spec_params:
            try:
                volume = float(volume_match.group(1))
                volume_widget = self.spec_params['volume'].get('widget')
                if volume_widget:
                    volume_widget.setValue(volume)
            except:
                pass
    
    def get_equipment(self):
        """获取设备对象 - 生成完整的规格字符串"""
        from ..process_design_data import EquipmentItem
        
        equipment_id = self.equipment_id_input.text().strip()
        name = self.name_input.text().strip()
        description_en = self.description_en_input.text().strip()
        
        # 如果用户输入了英文描述，更新对照表
        if name and description_en and self.parent_widget and hasattr(self.parent_widget, 'data_manager'):
            self.parent_widget.data_manager.add_equipment_name_mapping(name, description_en)
        
        # 构建规格字符串
        specification_parts = []
        
        # 添加通用规格
        if self.model_input.text().strip():
            specification_parts.append(f"型号: {self.model_input.text().strip()}")
        
        # 根据设备类型添加特定规格
        equipment_type = self.type_combo.currentText()
        for field_name, field_info in self.spec_params.items():
            widget = field_info.get('widget')
            label = field_info.get('label')
            unit = field_info.get('unit', '')
            field_type = field_info.get('type', '')
            
            value = None
            if field_type == 'spinbox':
                value = widget.value()
                if value > 0:
                    specification_parts.append(f"{label}: {value} {unit}")
            elif field_type == 'lineedit':
                value = widget.text().strip()
                if value:
                    specification_parts.append(f"{label}: {value}")
            elif field_type == 'checkbox':
                value = widget.isChecked()
                if value:
                    specification_parts.append(f"{label}: 是")
                else:
                    specification_parts.append(f"{label}: 否")
            elif field_type == 'textedit':
                value = widget.toPlainText().strip()
                if value:
                    # 其他规格可能会包含多条信息，每行作为单独的规格项
                    lines = value.split('\n')
                    for line in lines:
                        if line.strip():
                            specification_parts.append(line.strip())
        
        # 合并规格字符串
        specification = "; ".join(specification_parts)
        
        # 获取温度和压力值（如果勾选了常温常压，则设为"NT"和"NP"字符串）
        design_pressure = self.design_pressure_input.value() if self.design_pressure_input.isEnabled() else "NP"
        design_temperature = self.design_temperature_input.value() if self.design_temperature_input.isEnabled() else "NT"
        
        operating_pressure = self.operating_pressure_input.value() if self.operating_pressure_input.isEnabled() else "NP"
        operating_temperature = self.operating_temperature_input.value() if self.operating_temperature_input.isEnabled() else "NT"
        
        # 创建设备对象
        equipment_data = {
            'equipment_id': equipment_id,
            'name': name,
            'type': equipment_type,
            'model': self.model_input.text().strip(),
            'specification': specification,  # 完整的规格字符串
            'manufacturer': self.manufacturer_input.text().strip(),
            'location': self.location_input.text().strip(),
            'status': self.status_combo.currentText(),
            'commission_date': self.commission_date.text().strip() or None,
            'design_pressure': design_pressure,
            'design_temperature': design_temperature,
            'operating_pressure': operating_pressure,
            'operating_temperature': operating_temperature,
            'description': self.notes_input.toPlainText().strip(),
            'description_en': description_en,
            'pid_dwg_no': self.pid_dwg_no_input.text().strip(),
            'quantity': self.quantity_input.value(),
            'running_quantity': self.running_quantity_input.value(),
            'single_power': self.single_power_input.value() if self.single_power_input.value() > 0 else None,
            'operating_power': self.operating_power_input.value() if self.operating_power_input.value() > 0 else None,
            'total_power': self.total_power_input.value() if self.total_power_input.value() > 0 else None,
            'frequency_conversion': self.frequency_conversion_check.isChecked(),
            'unit_price': self.unit_price_input.value() if self.unit_price_input.value() > 0 else None,
            'total_price': self.total_price_input.value() if self.total_price_input.value() > 0 else None,
            'material': self.material_input.text().strip(),
            'insulation': self.insulation_input.text().strip(),
            'weight_estimate': self.weight_estimate_input.value() if self.weight_estimate_input.value() > 0 else None,
            'dynamic': self.dynamic_input.text().strip()
        }
        
        # 移除空字符串
        for key in ['material', 'insulation', 'dynamic']:
            if equipment_data[key] == '':
                equipment_data[key] = None
        
        # 创建设备对象
        equipment = EquipmentItem(**equipment_data)
        
        return equipment